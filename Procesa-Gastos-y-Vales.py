import openpyxl
import glob
import os


movimientos_inversiones = ['A', 'AS', 'TA', 'TE', 'TI']

excel_con_datos = glob.glob('*.xlsx')
contador = 0
for libro in excel_con_datos:
    if 'Gastos' in libro:
        contador += 1
        gastos = libro
if contador == 1:
    print('Se procesará el siguiente archivo de Gastos:', gastos)
    wb = openpyxl.load_workbook(gastos)
    sheet = wb.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    #Obtenemos el indice de la fila que contiene los nombres de las columnas
    for i in range(1, max_row + 1):
        if sheet.cell(row=i, column=max_column).value != None:
            indice_de_titulos = i
            break
    nombres_de_columnas = []
    for j in range(1, max_column + 1):
        nombres_de_columnas.append(sheet.cell(row=indice_de_titulos, column=j).value)
    for indice, columna in enumerate(nombres_de_columnas):
        if 'Accion' in columna:
            indice_columna_accion = indice + 1
        if 'INV' in columna:
            indice_columna_inversion = indice + 1
        if 'Servicio' in columna:
            indice_columna_servicio = indice + 1
    sheet.cell(row=indice_de_titulos, column=indice_columna_inversion + 1).value = 'Tipo_de_Movimiento'
    sheet.cell(row=indice_de_titulos, column=indice_columna_inversion + 2).value = 'DATOS'
    for i in range(1, max_row + 1):
        if sheet.cell(row=i, column=indice_columna_inversion).value:
            if i > 1:
                if sheet.cell(row=i, column=indice_columna_accion).value not in movimientos_inversiones:
                    intruso = sheet.cell(row=i, column=indice_columna_accion).value
                    print('ADVERTENCIA: Existe un movimiento de:', intruso, 'contemplado como Inversiones')
                if 'ED' in sheet.cell(row=i, column=indice_columna_servicio).value:
                    if sheet.cell(row=i, column=indice_columna_servicio + 1).value == 'MODEM':
                        sheet.cell(row=i, column=indice_columna_inversion + 2).value = 'Conectividad'
                        sheet.cell(row=i, column=indice_columna_inversion + 1).value = 'Inversiones'
                    else:
                        sheet.cell(row=i, column=indice_columna_inversion + 1).value = 'Preventivo'
                else:
                    sheet.cell(row=i, column=indice_columna_inversion + 1).value = 'Inversiones'
        else:
            if sheet.cell(row=i, column=indice_columna_accion).value != 'MttoP':
                sheet.cell(row=i, column=indice_columna_inversion + 1).value = 'Correctivo'
            else:
                sheet.cell(row=i, column=indice_columna_inversion + 1).value = 'Preventivo'
    wb.save(gastos)
    wb.close()

archivo = 'Materiales_abreviados.txt'
materiales = {}

if os.path.isfile(archivo):
    lectura_del_txt = open(archivo, "r", encoding="latin-1").readlines()
else:
    print("IMPORTANTE: El fichero 'Materiales_abreviados.txt' no se encuentra en la carpeta, este fichero es imprescindible para el funcionamiento del script")
    quit()
for linea in lectura_del_txt:
    codigo = linea.split(',')[0].strip()
    descripcion = linea.split(',')[1].strip()
    if descripcion:
        materiales.update({codigo: descripcion})

contador = 0
for libro in excel_con_datos:
    if 'Vales' in libro:
        contador += 1
        vales = libro
if contador == 1:
    print('Se procesará el siguiente archivo de Vales:', vales)
    wb = openpyxl.load_workbook(vales)
    sheet = wb.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    #Obtenemos el indice de la fila que contiene los nombres de las columnas
    for i in range(1, max_row + 1):
        if sheet.cell(row=i, column=max_column).value != None:
            indice_de_titulos = i
            break
    nombres_de_columnas = []
    for j in range(1, max_column + 1):
        nombres_de_columnas.append(sheet.cell(row=indice_de_titulos, column=j).value)
    ultima_columna = len(nombres_de_columnas) + 1
    sheet.cell(row=indice_de_titulos, column=ultima_columna).value = 'Descriptivo'
    for indice, columna in enumerate(nombres_de_columnas):
        if 'Material' in columna:
            indice_columna_material = indice + 1
    for i in range(1, max_row + 1):
        material_a_buscar = sheet.cell(row=i, column=indice_columna_material).value
        if str(material_a_buscar) in materiales:
            escribir = materiales[str(material_a_buscar)]
            sheet.cell(row=i, column=ultima_columna).value = escribir
    wb.save(vales)
    wb.close()


