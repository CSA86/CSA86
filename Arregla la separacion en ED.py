import openpyxl
import re

#file_path = 'SIPREC.xlsx'
file_path ='DatosReporteComercial.xlsx'

wb = openpyxl.load_workbook(file_path)
sheet = wb.active
max_row = sheet.max_row
max_column = sheet.max_column

for i in range(1, max_column + 1):
    nombre_del_campo = sheet.cell(row=1, column=i).value
    if nombre_del_campo =='SERVICIO' or nombre_del_campo == 'Servicio':
        indice_de_columna_servicio = i

for j in range(2,max_row+1):
    valor_del_campo = sheet.cell(row=j, column=indice_de_columna_servicio).value
    if valor_del_campo is not None and valor_del_campo is not "":
        mitad1 = valor_del_campo.split()[0]
        mitad2 = valor_del_campo.split()[1]
        valor_del_campo = mitad1 + mitad2
        print(valor_del_campo)
        sheet.cell(row=j, column=indice_de_columna_servicio).value = valor_del_campo

wb.save(file_path)