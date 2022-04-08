import openpyxl
import re
import glob
import os

estado = 'PENDIENTE POR PIEZAS RECOGIDAS'
nombre_de_piezas_por_modelo = 'Piezas_por_modelo.txt'
nombre_centros = 'Centros.txt'
piezas_por_modelo = {}
todos_los_centros = []


def agrupa_las_ep_por_centro(centros, listado_ep, piezas_por_modelo):
    print('RESUMEN PENDIENTE DE PIEZAS PARA LA DT')
    print('----------------------------------------------------------------------')
    llamadas_a_la_funcion = 0
    agrupa_las_ep_por_modelo('provincial', listado_ep, piezas_por_modelo, llamadas_a_la_funcion)
    print('\n')
    print('RESUMEN PENDIENTE DE PIEZAS POR MUNICIPIOS')
    for centro in centros:
        print('----------------------------------------------------------------------')
        print(centro)
        por_municipio = []
        for ep in listado_ep:
            if ep[4] in centro:
                por_municipio.append(ep)
        llamadas_a_la_funcion += 1
        agrupa_las_ep_por_modelo(centro, por_municipio, piezas_por_modelo, llamadas_a_la_funcion)

def agrupa_las_ep_por_modelo(centro, listado, piezas_x_modelo, llamadas_a_abre_excel):
    por_modelo = []
    for modelo in piezas_x_modelo.keys():
        for ep in listado:
            if modelo == ep[1]:
                por_modelo.append(ep)
    for modelo, piezas in piezas_x_modelo.items():
        determinante = 0
        conteo_por_modelo = {}
        conteo_por_pieza = {}
        for ep in por_modelo:
            if modelo == ep[1]:
                if modelo not in conteo_por_modelo:
                    conteo_por_modelo.update({modelo: 1})
                else:
                    valor_actual = conteo_por_modelo[modelo]
                    valor_actual += 1
                    conteo_por_modelo.update({modelo: valor_actual})
                for pieza in piezas:
                    if pieza in ep[2]:
                        if pieza not in conteo_por_pieza:
                            conteo_por_pieza.update({pieza: 1})
                        else:
                            valor_actual = conteo_por_pieza[pieza]
                            valor_actual += 1
                            conteo_por_pieza.update({pieza: valor_actual})
        if len(conteo_por_modelo) > 0:
            for clave in conteo_por_modelo.keys():
                modelo_a_imprimir = clave
            if len(conteo_por_pieza) > 0:
                llamadas_a_abre_excel += 1
                abre_y_escribe_en_excel(llamadas_a_abre_excel, centro, conteo_por_modelo, conteo_por_pieza)

def abre_y_escribe_en_excel(llamadas, centro, conteo_por_modelo, conteo_por_pieza):
    excel_con_datos = glob.glob('*.xlsx')
    contador = 0
    for libro in excel_con_datos:
        if 'Resumen' in libro:
            nombre_del_libro = libro
            contador += 1
    if contador == 1:
        file_path = nombre_del_libro
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        max_column = sheet.max_column
        max_row = sheet.max_row
        if llamadas == 1:
            sheet.delete_cols(1, 8)
            fila_provincial = 0
        else:
            fila_provincial = max_row
        if centro == 'provincial':
            for clave in conteo_por_modelo.keys():
                modelo_a_imprimir = clave
                for pieza, cantidad in conteo_por_pieza.items():
                    fila_provincial += 1
                    if fila_provincial > 1:
                        sheet.cell(row=fila_provincial, column=1).value = modelo_a_imprimir
                        sheet.cell(row=fila_provincial, column=2).value = pieza
                        sheet.cell(row=fila_provincial, column=3).value = cantidad
                    else:
                        sheet.cell(row=fila_provincial, column=1).value = "Modelo"
                        sheet.cell(row=fila_provincial, column=2).value = "Pieza"
                        sheet.cell(row=fila_provincial, column=3).value = "Cantidad"
        else:
            if sheet.cell(row=1, column=5).value is None:
                fila_por_municipio = 0
            else:
                for i in range(1, 700):
                    if sheet.cell(row=i, column=5).value is None:
                        fila_por_municipio = i - 1
                        break
            for clave in conteo_por_modelo.keys():
                modelo_a_imprimir = clave
                for pieza, cantidad in conteo_por_pieza.items():
                    fila_por_municipio += 1
                    if fila_por_municipio > 1:
                        sheet.cell(row=fila_por_municipio, column=5).value = centro
                        sheet.cell(row=fila_por_municipio, column=6).value = modelo_a_imprimir
                        sheet.cell(row=fila_por_municipio, column=7).value = pieza
                        sheet.cell(row=fila_por_municipio, column=8).value = cantidad
                    else:
                        sheet.cell(row=fila_por_municipio, column=5).value = "Centro"
                        sheet.cell(row=fila_por_municipio, column=6).value = "Modelo"
                        sheet.cell(row=fila_por_municipio, column=7).value = "Pieza"
                        sheet.cell(row=fila_por_municipio, column=8).value = "Cantidad"
    else:
            print('Deje solo un archivo "Resumen..." en la carpeta y vuelva a ejecutar el script')
            quit()
    wb.save(file_path)
    wb.close()

#Carga de un txt en la carpeta una lista con todos_los_centros
if os.path.isfile(nombre_centros):
    f_logfile = open(nombre_centros, "r", encoding="latin-1").readlines()
else:
    print("IMPORTANTE: El fichero 'Centros.txt' no se encuentra en la carpeta, este fichero es imprescindible para el funcionamiento del script")
    quit()
for linea in f_logfile:
    todos_los_centros.append(linea.strip())

#Carga de un txt en la carpeta el diccionario piezas_por_modelo
if os.path.isfile(nombre_de_piezas_por_modelo):
    f_logfile = open(nombre_de_piezas_por_modelo, "r", encoding="latin-1").readlines()
else:
    print("IMPORTANTE: El fichero 'Piezas_por_modelo.txt' no se encuentra en la carpeta, este fichero es imprescindible para el funcionamiento del script")
    quit()
for linea in f_logfile:
    modelo = linea.split(',')[0].strip()
    pieza = linea.split(',')[-1].strip()
    if modelo not in piezas_por_modelo:
        piezas_por_modelo.update({modelo:[pieza]})
    else:
        piezas_a_actualizar = piezas_por_modelo[modelo]
        piezas_a_actualizar.append(pieza)
        piezas_por_modelo.update({modelo: piezas_a_actualizar})

#Carga el excel del Total_de_Averias
excel_con_datos = glob.glob('*.xlsx')
contador = 0
for libro in excel_con_datos:
    if 'Total_de_Averías' in libro:
        nombre_del_libro = libro
        contador += 1
if contador == 1:
    print('El siguiente documento es el que se procesará para la distribución de las piezas de las EP interrumpidas:', nombre_del_libro)
    file_path = libro
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    #Obtenemos el indice de la fila que contiene los nombres de las columnas
    for i in range(1, max_row):
        valor_celda = sheet.cell(row=i, column=max_column).value
        if valor_celda != None:
            indice_de_titulos = i
            break
    nombres_de_columnas = []
    for j in range(1, max_column):
        nombres_de_columnas.append(sheet.cell(row=indice_de_titulos, column=j).value)
    for indice, columna in enumerate(nombres_de_columnas):
        if 'estado' in columna:
            indice_columna_estado = indice + 1
        elif 'mero' in columna:
            indice_columna_numero = indice + 1
        elif 'pieza' in columna:
            indice_columna_pieza = indice + 1
        elif 'modelo' in columna:
            indice_columna_modelo = indice + 1
        elif 'entro' in columna:
            indice_columna_centro = indice + 1
        elif 'ategor' in columna:
            indice_columna_categoria = indice + 1
        elif 'direc' in columna:
            indice_columna_direccion = indice + 1
        elif 'zona' in columna:
            indice_columna_zona = indice + 1
    listado_ep_pend_piezas = []
    for i in range(indice_de_titulos, max_row + 1):
        if sheet.cell(row=i, column=indice_columna_estado).value == estado:
            numero = sheet.cell(row=i, column=indice_columna_numero).value
            modelo = sheet.cell(row=i, column=indice_columna_modelo).value
            piezas_pendientes = sheet.cell(row=i, column=indice_columna_pieza).value
            categoria = sheet.cell(row=i, column=indice_columna_categoria).value
            centro_ep = sheet.cell(row=i, column=indice_columna_centro).value
            direccion = sheet.cell(row=i, column=indice_columna_direccion).value
            zona = sheet.cell(row=i, column=indice_columna_zona).value
            temporal = [numero, modelo, piezas_pendientes, categoria, centro_ep, direccion, zona]
            listado_ep_pend_piezas.append(temporal)
    agrupa_las_ep_por_centro(todos_los_centros, listado_ep_pend_piezas, piezas_por_modelo)
else:
    print('Revise la existencia de un solo documento excel con nombre Total_de_Averías')
    quit()