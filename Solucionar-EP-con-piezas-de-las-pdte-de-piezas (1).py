import openpyxl
import glob
import os

piezas_por_modelo = {}
todos_los_centros = []
centros_principales = {}
estado = 'PENDIENTE POR PIEZAS RECOGIDAS'

nombre_de_piezas_por_modelo = 'Piezas_por_modelo.txt'
nombre_centros = 'Centros.txt'
nombre_por_centros_principales = 'Centros_agrupados_por_CTP.txt'

def analisis(listado_ep, ent, opc, ct, ctp, piezas_x_modelos, por_zona):
    if ent == 'a' or ent == 'A':
        if opc == '0':
            for centro_escogido in ct:
                agrupa_las_ep_por_centro(centro_escogido, listado_ep, piezas_x_modelos)
        else:
            centro_escogido = ct[int(opc) - 1]
            if centro_escogido == 'HOLGUIN' and por_zona == '0':
                agrupa_las_ep_por_zona(listado_ep, piezas_x_modelos)
            else:
                agrupa_las_ep_por_centro(centro_escogido, listado_ep, piezas_x_modelos)
    elif ent == 'b' or ent == 'B':
        if opc == '0':
            for centro_escogido in ctp.values():
                agrupa_las_ep_por_centro(centro_escogido, listado_ep, piezas_x_modelos)
        else:
            orden = 0
            for clave in ctp.keys():
                orden += 1
                if orden == int(opc):
                    centro_escogido = ctp[clave]
            agrupa_las_ep_por_centro(centro_escogido, listado_ep, piezas_x_modelos)
    else:
        agrupa_las_ep_por_modelo(listado_ep, piezas_por_modelo)

def agrupa_las_ep_por_centro(centro, listado_ep,piezas_x_modelo):
    print("Se analizan las EP de:", centro)
    por_municipio = []
    for ep in listado_ep:
        if ep[4] in centro:
            por_municipio.append(ep)
    agrupa_las_ep_por_modelo(por_municipio, piezas_x_modelo)

def agrupa_las_ep_por_zona(listado_ep, piezas_x_modelo):
    zonas_ct_ho = []
    nombre_zonas = 'Zonas.txt'
    if os.path.isfile(nombre_zonas):
        f_logfile = open(nombre_zonas, "r", encoding="latin-1").readlines()
    else:
        print(
            "IMPORTANTE: El fichero 'Zonas.txt' no se encuentra en la carpeta, este fichero es imprescindible para el funcionamiento del script")
        quit()
    for linea in f_logfile:
        zonas_ct_ho.append(linea.strip())
    print("Se analizan las EP del municipio HOLGUIN por Zonas:")
    for zona in zonas_ct_ho:
        por_zonas = []
        print('******************************Zona:', zona, '******************************')
        for ep in listado_ep:
            if ep[6] == zona:
                por_zonas.append(ep)
        agrupa_las_ep_por_modelo(por_zonas, piezas_x_modelo)

def agrupa_las_ep_por_modelo(listado, piezas_x_modelo):
    por_modelo = []
    for modelo in piezas_x_modelo.keys():
        for ep in listado:
            if modelo == ep[1]:
                por_modelo.append(ep)
    for modelo, piezas in piezas_x_modelo.items():
        determinante = 0
        conteo_por_modelo = {}
        conteo_por_pieza = {}
        for ep in por_modelo:
            if modelo == ep[1]:
                if modelo not in conteo_por_modelo:
                    conteo_por_modelo.update({modelo: 1})
                else:
                    valor_actual = conteo_por_modelo[modelo]
                    valor_actual += 1
                    conteo_por_modelo.update({modelo: valor_actual})
                for pieza in piezas:
                    if pieza in ep[2]:
                        if pieza not in conteo_por_pieza:
                            conteo_por_pieza.update({pieza: 1})
                        else:
                            valor_actual = conteo_por_pieza[pieza]
                            valor_actual += 1
                            conteo_por_pieza.update({pieza: valor_actual})
        for valor in conteo_por_pieza.values():
            if valor > determinante:
                determinante = valor
        if len(conteo_por_modelo) > 0 and len(conteo_por_pieza) > 0 and conteo_por_modelo[modelo] - determinante > 0:
            print(conteo_por_modelo)
            print(conteo_por_pieza)
            salvables = conteo_por_modelo[modelo] - determinante
            print("Salvables para este modelo:", salvables)
            opciones_de_eleccion = []
            elecciones = []
            for orden, ep in enumerate(por_modelo):
                if modelo == ep[1]:
                    print(orden,'-',ep[0], ep[2],'Categoria:', ep[3],ep[5])
                    opciones_de_eleccion.append(str(orden))
            print("De las EP 'Salvables para este modelo', a continuación seleccione las que desee reparar")
            print("Puede especificar todas las EP, algunas o ninguna. Teclee s cuando concluya su selección")
            while len(elecciones) < salvables:
                opcion = input("Indique su elección:\n")
                if opcion == 's':
                    break
                else:
                    while opcion not in opciones_de_eleccion:
                        opcion = input("Por favor seleccione una de las opciones posibles:\n")
                elecciones.append(opcion)
            agrupa_en_a_sacrificar_y_a_reparar(por_modelo, opciones_de_eleccion, elecciones, piezas_x_modelo, salvables, modelo)
            print('-----------------------------------------------------------------------------')

def agrupa_en_a_sacrificar_y_a_reparar(por_modelo, opciones, elecciones, piezas_x_modelo, salvables, modelo):
    ep_a_reparar = []
    ep_a_sacrificar = []
    for orden, ep in enumerate(por_modelo):
        if ep[1] == modelo:
            ep_a_sacrificar.append(ep)
    if len(elecciones) < salvables:
        if len(elecciones) > 0:
            for salv in elecciones:
                temp = por_modelo[int(salv)]
                indice = ep_a_sacrificar.index(temp)
                ep_a_reparar.append(ep_a_sacrificar.pop(indice))
        escoger = salvables - len(elecciones)
        while escoger > 0:
            for orden, ep in enumerate(ep_a_sacrificar):
                if ep[3] == 1 and escoger > 0:
                    ep_a_reparar.append(ep_a_sacrificar.pop(orden))
                    escoger -= 1
            if escoger > 0:
                for orden, tp in enumerate(ep_a_sacrificar):
                    if 'EQUIPO COMPLETO' not in tp[2]:
                        ep_a_reparar.append(ep_a_sacrificar.pop(orden))
                        escoger -= 1
                        break
    else:
        for salvable in elecciones:
            temp = por_modelo[int(salvable)]
            indice = ep_a_sacrificar.index(temp)
            ep_a_reparar.append(ep_a_sacrificar.pop(indice))
    realiza_asignacion_de_piezas(ep_a_sacrificar, ep_a_reparar, piezas_x_modelo, modelo)

def realiza_asignacion_de_piezas(ep_a_sacrificar, ep_a_reparar, piezas_x_modelo, modelo):
    piezas_del_modelo = piezas_x_modelo[modelo]
    todas = set(piezas_del_modelo)
    for tp in ep_a_sacrificar:
        piezas_pendientes = convierte_piezas_pendientes_en_lista(tp)
        pendientes = set(piezas_pendientes)
        usables = todas.difference(pendientes)
        tp.append(list(usables))
    for ep in ep_a_reparar:
        piezas_a_reparar = convierte_piezas_pendientes_en_lista(ep)
        for ord, pieza in enumerate(piezas_a_reparar):
            encontrada = False
            if pieza != 'EQUIPO COMPLETO' and not encontrada:
                for tp in ep_a_sacrificar:
                    disponibles = tp[7]
                    if not encontrada:
                        for orden, disp in enumerate(disponibles):
                            if pieza == disp:
                                actualizar = pieza + ':' + str(tp[0])
                                ep.append(actualizar)
                                disponibles.pop(orden)
                                ep[2] = piezas_a_reparar
                                actuales = tp[2]
                                actuales = actuales + ', ' + disp
                                tp[2] = actuales
                                encontrada = True
    print('REPARADA:')
    for ep in ep_a_reparar:
        print(ep[0:2], ep[7:])
    print('PENDIENTE')
    for tp in ep_a_sacrificar:
        print(tp[0:3])

def convierte_piezas_pendientes_en_lista(tp):
    piezas = []
    piezas_pendientes = []
    piezas.append(tp[2])
    for pieza in piezas:
        otra_lista = pieza.split(', ')
        for caso in otra_lista:
            caso.strip()
            piezas_pendientes.append(caso)
    return piezas_pendientes



#Carga de un txt en la carpeta el diccionario piezas_por_modelo
if os.path.isfile(nombre_de_piezas_por_modelo):
    f_logfile = open(nombre_de_piezas_por_modelo, "r", encoding="latin-1").readlines()
else:
    print("IMPORTANTE: El fichero 'Piezas_por_modelo.txt' no se encuentra en la carpeta, este fichero es imprescindible para el funcionamiento del script")
    quit()
for linea in f_logfile:
    modelo = linea.split(',')[0].strip()
    pieza = linea.split(',')[-1].strip()
    if modelo not in piezas_por_modelo:
        piezas_por_modelo.update({modelo:[pieza]})
    else:
        piezas_a_actualizar = piezas_por_modelo[modelo]
        piezas_a_actualizar.append(pieza)
        piezas_por_modelo.update({modelo: piezas_a_actualizar})

#Carga de un txt en la carpeta una lista con todos_los_centros
if os.path.isfile(nombre_centros):
    f_logfile = open(nombre_centros, "r", encoding="latin-1").readlines()
else:
    print("IMPORTANTE: El fichero 'Centros.txt' no se encuentra en la carpeta, este fichero es imprescindible para el funcionamiento del script")
    quit()
for linea in f_logfile:
    todos_los_centros.append(linea.strip())

#Carga de un txt en la carpeta el diccionario para agrupar los centros por CTP
if os.path.isfile(nombre_por_centros_principales):
    f_logfile = open(nombre_por_centros_principales, "r", encoding="latin-1").readlines()
else:
    print("IMPORTANTE: El fichero 'Centros_agrupados_por_CTP.txt' no se encuentra en la carpeta, este fichero es imprescindible para el funcionamiento del script")
    quit()
for linea in f_logfile:
    ctp = linea.split(",")[0].strip()
    toda_la_linea = linea.split(',')
    if '\n' in toda_la_linea[-1]:
        toda_la_linea[-1] = toda_la_linea[-1].replace('\n','')
    centros_principales.update({ctp: toda_la_linea})

excel_con_datos = glob.glob('*.xlsx')
contador = 0
for libro in excel_con_datos:
    if 'Total_de_Averías' in libro:
        contador += 1
if contador == 1:
    print('El siguiente documento es el que se procesará para la distribución de las piezas de las EP interrumpidas:', libro)
    file_path = libro
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    #Obtenemos el indice de la fila que contiene los nombres de las columnas
    for i in range(1, max_row):
        if sheet.cell(row=i, column=max_column).value != None:
            indice_de_titulos = i
            break
    nombres_de_columnas = []
    for j in range(1, max_column):
        nombres_de_columnas.append(sheet.cell(row=indice_de_titulos, column=j).value)

    print("Indique con una letra, según corresponda, a que nivel se realizará la redistribución de las piezas de las EP interrumpidas:")
    print("A: Centro de Telecomunincaciones")
    print("B: Centro de Telecomunicaciones Principal")
    print("C: Provincial")
    entrada = input("Indique su elección:\n")
    opciones = ["A", "B", "C", "a", "b", "c"]
    while entrada not in opciones:
        entrada = input("Especifique con una letra alguna de las entradas válidas de las propuestas anteriormente:\n")
    if entrada == 'A' or entrada == 'a':
        print("Escoja el Centro de Telecomunicaciones, especificando el numero de orden:")
        print("0 - TODOS LOS CT")
        zona_o_municipio = '1'
        for orden, centro in enumerate(todos_los_centros):
            print(orden + 1, '-', centro)
        opciones = [str(i)for i in range(0, len(todos_los_centros) + 1)]
        opcion = input("Indique su elección:\n")
        while opcion not in opciones:
            opcion = input("Indique su elección:\n")
        if opcion == '1':
            print('Ha escogido el CT Holguin, especifique si desea una distribucion por Zona o a nivel de municipio:')
            print('0 - Zona')
            print('1 - Municipio')
            zona_o_municipio = input("Indique su elección:\n")
            opcion_zona_o_municipio = ['0', '1']
            while zona_o_municipio not in opcion_zona_o_municipio:
                zona_o_municipio = input("Indique su elección:\n")
    elif entrada == 'B' or entrada == 'b':
        print("Escoja el Centro de Telecomunicaciones Principal, especificando el numero de orden:")
        print("0 - TODOS LOS CT")
        orden = 0
        for centro in centros_principales.keys():
            orden += 1
            print(orden, '-', centro)
        opciones = [str(i) for i in range(0, len(centros_principales) + 1)]
        opcion = input("Indique su elección:\n")
        while opcion not in opciones:
            opcion = input("Indique su elección:\n")
    elif entrada == 'c' or entrada == 'C':
        print("Ha escogido realizar la distribución a nivel provincial")
        opcion = '0'
    for indice, columna in enumerate(nombres_de_columnas):
        if 'estado' in columna:
            indice_columna_estado = indice + 1
        elif 'mero' in columna:
            indice_columna_numero = indice + 1
        elif 'pieza' in columna:
            indice_columna_pieza = indice + 1
        elif 'modelo' in columna:
            indice_columna_modelo = indice + 1
        elif 'entro' in columna:
            indice_columna_centro = indice + 1
        elif 'ategor' in columna:
            indice_columna_categoria = indice + 1
        elif 'direc' in columna:
            indice_columna_direccion = indice + 1
        elif 'zona' in columna:
            indice_columna_zona = indice + 1
    listado_ep_pend_piezas = []
    for i in range(indice_de_titulos, max_row + 1):
        if sheet.cell(row=i, column=indice_columna_estado).value == estado:
            numero = sheet.cell(row=i, column=indice_columna_numero).value
            modelo = sheet.cell(row=i, column=indice_columna_modelo).value
            piezas_pendientes = sheet.cell(row=i, column=indice_columna_pieza).value
            categoria = sheet.cell(row=i, column=indice_columna_categoria).value
            centro_ep = sheet.cell(row=i, column=indice_columna_centro).value
            direccion = sheet.cell(row=i, column=indice_columna_direccion).value
            zona = sheet.cell(row=i, column=indice_columna_zona).value
            temporal = [numero, modelo, piezas_pendientes, categoria, centro_ep, direccion, zona]
            listado_ep_pend_piezas.append(temporal)
    analisis(listado_ep_pend_piezas, entrada, opcion, todos_los_centros, centros_principales, piezas_por_modelo, zona_o_municipio)
else:
    print('Revise la existencia de un solo documento excel con nombre Total_de_Averías')








