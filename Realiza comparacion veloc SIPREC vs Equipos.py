import openpyxl
import re

def prepara_valor_siprec(speed):
    velocidades = []
    if '/' in speed:
        descarga = int(speed.split('/')[0].strip())
        subida = int(speed.split('/')[1].strip())
        descarga = int(descarga/1000)
        subida = int(subida/1000)
    else:
        descarga = int(int(speed)/1000)
        subida = descarga
    velocidades.append(descarga)
    velocidades.append(subida)
    return velocidades

def prepara_valor_equipo(velocidad):
    speed = 0
    if velocidad is not None:
        if int(velocidad) < 128:
            speed = 64
        elif int(velocidad) >= 128 and int(velocidad) < 256:
            speed = 128
        elif int(velocidad) >= 256 and int(velocidad) < 512:
            speed = 256
        elif int(velocidad) >= 512 and int(velocidad) < 1024:
            speed = 512
        elif int(velocidad) >= 1024 and int(velocidad) < 1536:
            speed = 1024
        elif int(velocidad) >= 1536 and int(velocidad) < 2048:
            speed = 1536
        elif int(velocidad) >= 2048 and int(velocidad) < 4096:
            speed = 2048
        elif int(velocidad) >= 4096 and int(velocidad) < 6144:
            speed = 4096
        elif int(velocidad) >= 6144 and int(velocidad) < 8192:
            speed = 6144
        elif int(velocidad) >= 8192 and int(velocidad) < 10240:
            speed = 8192
        elif int(velocidad) >= 10240 and int(velocidad) < 20480:
            speed = 10240
        elif int(velocidad) >= 20480:
            speed = 20480
    else:
        speed = 0
    return speed

def comparacion_de_valores(veloc_siprec, desc_linea, sub_linea, desc_traf, sub_traf):
    respuesta = True
    descarga_equipo = 0
    subida_equipo = 0
    if desc_linea > 0 and sub_linea > 0 and desc_traf > 0 and sub_traf > 0:
        if desc_linea <= desc_traf:
            descarga_equipo = desc_linea
        else:
            descarga_equipo = desc_traf
        if sub_linea <= sub_traf:
            subida_equipo = sub_linea
        else:
            subida_equipo = sub_traf
    else:
        if desc_linea == 0 and desc_traf > 0:
            descarga_equipo = desc_traf
        elif desc_linea > 0 and desc_traf == 0:
            descarga_equipo = desc_linea
        elif desc_linea == 0 and desc_traf == 0:
            descarga_equipo = 0
        if sub_linea == 0 and sub_traf > 0:
            subida_equipo = sub_traf
        elif sub_linea > 0 and sub_traf == 0:
            subida_equipo = sub_linea
        elif sub_linea == 0 and sub_traf == 0:
            subida_equipo == 0
    if veloc_siprec[0] != descarga_equipo:
        respuesta = False
    if veloc_siprec[1] != subida_equipo:
        respuesta = False
    return respuesta



valores_siprec = []
file_path = '1-Están en Comercial y en Equipo compara velocidad.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active
max_row = sheet.max_row
max_column = sheet.max_column

for i in range(1, max_column + 1):
    nombre_del_campo = sheet.cell(row = 1, column = i).value
    if nombre_del_campo == 'Velocidad':
        indice_de_columna_siprec = i
    elif nombre_del_campo == 'Veloc_descarga_perfil_linea':
        indice_de_columna_descarga_perfil_linea = i
    elif nombre_del_campo == 'Veloc_subida_perfil_linea':
        indice_de_columna_subida_perfil_linea = i
    elif nombre_del_campo == 'Veloc_descarga_perfil_trafico':
        indice_de_columna_descarga_perfil_trafico = i
    elif nombre_del_campo == 'Veloc_subida_perfil_trafico':
        indice_de_columna_subida_perfil_trafico = i

for j in range(2,max_row+1):
    siprec = sheet.cell(row = j, column = indice_de_columna_siprec).value
    valores_siprec = prepara_valor_siprec(siprec)
    descarga_perfil_linea = sheet.cell(row = j, column = indice_de_columna_descarga_perfil_linea).value
    descarga_linea = prepara_valor_equipo(descarga_perfil_linea)
    subida_perfil_linea = sheet.cell(row = j, column = indice_de_columna_subida_perfil_linea).value
    subida_linea = prepara_valor_equipo(subida_perfil_linea)
    descarga_perfil_trafico = sheet.cell(row = j, column = indice_de_columna_descarga_perfil_trafico).value
    descarga_trafico = prepara_valor_equipo(descarga_perfil_trafico)
    subida_perfil_trafico = sheet.cell(row = j, column = indice_de_columna_subida_perfil_trafico).value
    subida_trafico = prepara_valor_equipo(subida_perfil_trafico)
    answer = comparacion_de_valores(valores_siprec, descarga_linea, subida_linea, descarga_trafico, subida_trafico)
    if answer == False:
        sheet.cell(row = j, column = max_column + 1).value = answer
wb.save(file_path)
wb.close