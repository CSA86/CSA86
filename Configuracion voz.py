import os
import glob
import copy
import openpyxl
from time import time

start_time = time()

def analiza_interface_en_nanta(datos, vlan_voz):
    respuesta = {}
    existe_vlan_voz = ''
    ip_rtp = 'ip ' + str(vlan_voz)
    vlan = 'vlan-id ' + str(vlan_voz)
    vrf_id = ''
    ip3 = ''
    for orden, caso in enumerate(datos):
        if vlan in caso:
            existe_vlan_voz = datos[orden +1].strip()
        if ip_rtp in caso:
            for orden_siguiente, caso_siguiente in enumerate(datos[orden + 1:]):
                if 'exit' in caso_siguiente:
                    break
                else:
                    if 'vrf-id' in caso_siguiente:
                        vrf_id = caso_siguiente.strip()
                    if 'ip-addr' in caso_siguiente:
                        ip3 = caso_siguiente.split('ip-addr ')[-1].strip()
    if not existe_vlan_voz:
        print('ADVERTENCIA: No fue configurada la vlan de voz en la interface shub')
    else:
        if existe_vlan_voz != 'admin-status up':
            print('ADVERTENCIA: La vlan de voz en el interface shub no se encuentra admin-status up')
    if not vrf_id:
        print("ADVERTENCIA: No se especifico el vrf al que se asocia la interfaz ip de la vlan de rtp")
    if not ip3:
        print('ADVERTENCIA:No se especificó la IP3 en la interfaz ip de la vlan de RTP')
        return respuesta
    else:
        respuesta.update({"IP3": ip3})
        escribir = 'IP3:' + ip3 + '\n'
        resultado.write(escribir)
        return respuesta

def analiza_vrf(datos, vlan_voz):
    retorno_de_la_funcion = {}
    vlan = ''
    next_hop = ''
    for caso in datos:
        if 'vrf 1' in caso:
            if 'fast-path-mode:ena-user-user-com' not in caso:
                print('ADVERTENCIA: No se le configuró el parámetro fast-path-mode:ena-user-user-com al vrf')
        if 'route-dest 0.0.0.0/0' in caso:
            if 'next-hop' in caso:
                next_hop = caso.split('next-hop ')[-1].split()[0].strip()
            else:
                print('ADVERTENCIA: No se configuró el proximo salto dentro del vrf')
            if 'vlan-id' in caso:
                vlan = caso.split('vlan-id ')[-1].strip()
            else:
                print("ADVERTENCIA: No se configuró vlan de voz dentro del vrf")
    if vlan_voz != vlan:
        print('ADVERTENCIA: La vlan configurada en el vrf', vlan, 'no coincide con la vlan de rtp del equipo', vlan_voz)
    if next_hop:
        retorno_de_la_funcion.update({'Default Router RTP':next_hop})
        return retorno_de_la_funcion
    else:
        return retorno_de_la_funcion


def devuelve_ip_gestion_en_nanta(datos):
    respuesta = {}
    ip1 = ''
    default_route = ''
    for caso in datos:
        if 'host-ip-address manual:' in caso:
            ip1 = caso.split('host-ip-address manual:')[-1].strip()
        if 'default-route' in caso:
            default_route = caso.split()[-1].strip()
    respuesta.update({'ip1': ip1})
    respuesta.update({'Default router gestion': default_route})
    escribir = 'IP1:' + ip1 + ' Default router: ' + default_route + '\n'
    resultado.write(escribir)
    return respuesta

def verifica_puerto_y_tarjetas_dentro_de_las_vlan_senalizacion_rtp_gestion_nant_a(datos_vlan, vlan_gestion, vlan_voz, vlan_senalizacion, tarj, sustendido):
    uplink_de_senalizacion = []
    uplink_de_gestion = []
    puertos_de_voz = []
    tarjetas_gestion = []
    tarjetas_voz = []
    tarjetas_nvps = []
    for orden, caso in enumerate(datos_vlan):
        if 'id' in caso:
            if vlan_gestion in caso:
                tipo_de_vlan = caso.split('id ')[-1].split('mode ')[-1].strip()
                if tipo_de_vlan != 'reserved':
                    print('ADVERTENCIA: La vlan de gestión no fue creada como mode reserved')
                else:
                    for orden1, caso1 in enumerate(datos_vlan[orden + 1:]):
                        if 'id' in caso1:
                            break
                        else:
                            if 'egress-port' in caso1:
                                tarjetas_gestion.append(caso1.split('egress-port ')[-1].strip())
            if vlan_voz in caso:
                tipo_de_vlan = caso.split('id ')[-1].split('mode ')[-1].strip()
                if tipo_de_vlan != 'voice-vlan':
                    print('ADVERTENCIA: La vlan de voz no fue creada como mode voice-vlan')
                else:
                    for orden1, caso1 in enumerate(datos_vlan[orden + 1:]):
                        if 'id' in caso1:
                            break
                        else:
                            if 'egress-port' in caso1:
                                tarjetas_voz.append(caso1.split('egress-port ')[-1].strip())
            if vlan_senalizacion in caso:
                tipo_de_vlan = caso.split('id ')[-1].split('mode ')[-1].strip()
                if tipo_de_vlan != 'residential-bridge':
                    print('ADVERTENCIA: La vlan de señalización no fue creada como mode residential-bridge')
                else:
                    for orden1, caso1 in enumerate(datos_vlan[orden + 1:]):
                        if 'id' in caso1:
                            break
                        else:
                            if 'egress-port' in caso1:
                                tarjetas_nvps.append(caso1.split('egress-port ')[-1].strip())
    for clave, valor in tarj.items():
        if 'npot' in valor or 'nvps' in valor:
            tarjetas_voz.remove(clave)
    if tarjetas_voz:
        for x in tarjetas_voz:
            if 'lt:' in x:
                print('ADVERTENCIA: Se configuró la siguiente tarjeta dentro de la vlan de rtp y la misma no es NVPS ni NPOT: ', x)
            if 'network:' in x:
                puertos_de_voz.append(x)
    if puertos_de_voz:
        if len(puertos_de_voz) < 2 and sustendido:
            print("ADVERTENCIA: Solo hay 1 puerto network configurado dentro de la vlan de voz y existe un ISAM Sustendido:", puertos_de_voz)
        if len(puertos_de_voz) == 2 and sustendido:
            print('Estos son los puertos network configurados en la vlan de voz, verifique en cada caso cual es el puerto de uplink y cual el del sustendido:', puertos_de_voz)
        if len(puertos_de_voz) > 1 and not sustendido:
            print("ADVERTENCIA: Existen puertos network que no es necesario que estén en la vlan de voz, verifique:", puertos_de_voz)
    else:
        print("ADVERTENCIA: No existen puertos network configurados dentro de la vlan de voz")
    for clave, valor in tarj.items():
        if 'nvps' in valor:
            if clave in tarjetas_nvps:
                tarjetas_nvps.remove(clave)
            else:
                print('ADVERTENCIA: La tarjeta NVPS ', clave, 'no se encuentra permitida en la vlan de señalización')
    if tarjetas_nvps:
        for x in tarjetas_nvps:
            if 'lt:' in x:
                print('ADVERTENCIA: Se configuró la siguiente tarjeta dentro de la vlan de señalización y la misma no es NVPS: ', x)
            if 'network:' in x:
                uplink_de_senalizacion.append(x)
                print('Verifique que este puerto de las NANT-A es el puerto de uplink del equipo:', x)
        if not uplink_de_senalizacion:
            print('ADVERTENCIA: La vlan de señalización no se encuentra permitida en ninguno de los puertos network')
        if len(uplink_de_senalizacion) > 1:
            print('ADVERTENCIA: Se configuró la vlan de señalización en más de 1 puerto, solo debe estar en el de uplink:', uplink_de_senalizacion)
    else:
        print('ADVERTENCIA: La vlan de señalización no se encuentra permitida en ninguno de los puertos network')
    if tarjetas_gestion:
        for clave, valor in tarj.items():
            if 'nvps' in valor:
                tarjetas_gestion.remove(clave)
    if tarjetas_gestion:
        for x in tarjetas_gestion:
            if 'lt:' in x:
                print("ADVERTENCIA: Se configuró la tarjeta ", x, 'dentro de la vlan de señalización sin ser necesario')
            if 'network:' in x:
                uplink_de_gestion.append(x)
        if not uplink_de_gestion:
            print('ADVERTENCIA: La vlan de gestión no se encuentra permitida en ninguno de los puertos network')
        if len(uplink_de_gestion) == 1 and sustendido:
            print("ADVERTENCIA: La vlan de gestión solo se encuentra permitida en un puerto network y existe un ISAM sustendido, verifique:", uplink_de_gestion[0])
        if len(uplink_de_gestion) == 2 and sustendido:
            print('Estos son los puertos network configurados en la vlan de gestión, verifique en cada caso cual es el puerto de uplink y cual el del sustendido:', uplink_de_gestion)
    else:
        print('ADVERTENCIA: La vlan de gestión no se encuentra permitida en ninguno de los puertos network')


def verifica_gestion_del_equipo(serv_configurado, gestion, card, link_aggregation):
    tarjetas_que_deben_tener_gestion = []
    vpls_de_gestion = []
    existe_descripcion = False
    ip_gestion = ''
    for clave, valor in serv_configurado.items():
        if 'ies' in clave:
            virtual_port = 'nt:vp:1:' + str(gestion)
            if str(virtual_port) in str(valor):
                indice_virtual_port = valor.index(virtual_port)
                for caso in valor:
                    if '/' in caso and '.' in caso:
                        if valor.index(caso) == indice_virtual_port - 1:
                            ip_gestion = caso
        if str(gestion) == str(clave.split(' customer')[0].split()[-1].strip()):
            vpls_de_gestion.append(valor)
            for parametro in valor:
                if '"' in parametro:
                    existe_descripcion = True
    for clave, valor in card.items():
        if 'nvps-c' in valor or 'fglt-b' in valor:
            tarjetas_que_deben_tener_gestion.append(clave + ':' + gestion)
    for clave in link_aggregation.keys():
        tarjetas_que_deben_tener_gestion.append('lag-' + clave.split()[-1].strip() + ':' + gestion)
    if vpls_de_gestion:
        for caso in tarjetas_que_deben_tener_gestion:
            if caso not in vpls_de_gestion[0]:
                if 'lt' in caso:
                    print('ADVERTENCIA: La tarjeta ubicada en la ranura', caso, 'debe estar permitida en la vpls de gestión')
                else:
                    print('ADVERTENCIA: El', caso, 'no se encuentra permitido dentro de la vpls de gestión')
    if not existe_descripcion:
        print('ADVERTENCIA: No se le especificó Descripción a la Vlan de Gestión')
    if not vpls_de_gestion:
        print("ADVERTENCIA: No se ha configurado la vlan de gestión o se encuentra shutdown")
    escribir = 'IP1: ' + ip_gestion + '\n'
    resultado.write(escribir)
    return ip_gestion

def analiza_puertos(informacion):
    respuesta = {}
    for orden, caso in enumerate(informacion):
        caracteristicas_del_puerto = []
        if "port" in caso:
            port = caso.split()[-1].strip()
            fin_del_caso = False
            next = orden
            while not fin_del_caso:
                if next + 1 >= len(informacion) or "port" in informacion[next + 1]:
                    fin_del_caso = True
                else:
                    next += 1
                    if "ethernet" not in informacion[next] and "exit" not in informacion[next]:
                        caracteristicas_del_puerto.append(informacion[next].strip())
                        if 'no shutdown' in caracteristicas_del_puerto:
                            respuesta.update({port: caracteristicas_del_puerto})
    return respuesta

def analiza_lag(informacion):
    respuesta = {}
    for orden, caso in enumerate(informacion):
        caracteristicas_del_lag = []
        if "lag" in caso:
            lag = caso.strip()
            fin_del_caso = False
            next = orden
            while not fin_del_caso:
                if next + 1 >= len(informacion) or "exit" in informacion[next + 1]:
                    fin_del_caso = True
                else:
                    next += 1
                    a_archivar = informacion[next].strip()
                    if "port " in a_archivar:
                        a_archivar = a_archivar.split("port ")[-1].strip()
                    caracteristicas_del_lag.append(a_archivar)
                    if 'no shutdown' in caracteristicas_del_lag:
                        respuesta.update({lag: caracteristicas_del_lag})
    return respuesta

def analiza_servicios(informacion, rtp_vlan, senalizacion_vlan, tarj, es_sustendido):
    respuesta = {}
    existe_vlan_para_rtp = False
    existe_vlan_senalizacion = False
    datos_configurados = []
    for orden, caso in enumerate(informacion):
        #print(caso)
        caracteristicas = []
        if "customer" in caso and not "description" in caso or "ies" in caso or "vprn" in caso or "vpls" in caso :
            #print(caso)
            if caso not in datos_configurados:
                datos_configurados.append(caso)
    for case in datos_configurados:
        if rtp_vlan in case:
            existe_vlan_para_rtp = True
        if senalizacion_vlan in case:
            existe_vlan_senalizacion = True
    if not existe_vlan_para_rtp:
        print("ALERTA: No existe VPLS para RTP")
    if not existe_vlan_senalizacion and not es_sustendido:
        print("ALERTA: No existe VPLS para SENALIZACION")
    for orden1, caso1 in enumerate(informacion):
        fin_del_caso = False
        next = orden1
        caracteristicas_del_servicio = []
        if caso1 in datos_configurados:
            while not fin_del_caso:
                if next + 1 >= len(informacion) or informacion[next + 1] in datos_configurados:
                    fin_del_caso = True
                else:
                    next += 1
                    x = informacion[next].strip()
                    if " create" in x:
                        x = x.split(" create")[0].strip()
                    if "sap " in x:
                        x = x.split("sap ")[-1].strip()
                    if "address " in x:
                        x = x.split("address ")[-1].strip()
                    if "description " in x:
                        x = x.split("description ")[-1].strip()
                    if "interface " in x:
                        x = x.split("interface ")[-1].strip()
                    if "exit" not in x and "stp" not in x and x != "shutdown" and x != "enable-stats":
                        if x not in caracteristicas_del_servicio:
                            caracteristicas_del_servicio.append(x)
            if "no shutdown" in caracteristicas_del_servicio[-1]:
                respuesta.update({caso1: caracteristicas_del_servicio})
    return respuesta

def verifica_puertos_del_lag_operativos(puertos_operativos, lag_existentes):
    for clave, valor in lag_existentes.items():
        for caso in valor:
            if 'description' not in caso and 'lacp' not in caso and 'no shutdown' not in caso:
                if caso not in puertos_operativos.keys():
                    print("ADVERTENCIA: El puerto",caso, "que esta dentro del", clave, "no se encuentra operativo")

def verifica_operativos_puertos_de_tarjetas_y_virtual_port(tarjetas_en_uso, puertos_en_uso):
    for clave in tarjetas_en_uso.keys():
        puerto_de_nant_enecendido = False
        if "nt-" not in clave:
            if clave not in puertos_en_uso.keys():
                print("ADVERTENCIA: El puerto de la tarjeta", clave, "esta apagado")
        else:
            for keys in puertos_en_uso.keys():
                if clave in keys:
                    puerto_de_nant_enecendido = True
            if not puerto_de_nant_enecendido:
                print("ADVERTENCIA: La tarjeta de control", clave, "no tiene encendido ningun un puerto")
    if 'nt:vp:1' not in puertos_en_uso:
        print("ADVERTENCIA: El virtual port (se usa en la VPRN 80008) se encuentra apagado")

def verifica_vpls_senalizacion_incluya_NVPS_LAG_puerto_uplink(tarjetas, puertos, lag, servicios, vlan_senalizacion):
    nvps = []
    para_encontrar_lag = encontrar_lag(lag)
    existe_vpls_senalizacion = False
    for clave, valor in tarjetas.items():
        if valor == 'nvps-c':
            nvps.append(clave + ":")
    for clave, valor in servicios.items():
        x = clave.split(" customer")[0].split("vpls ")[-1].strip()
        if vlan_senalizacion == x:
            existe_vpls_senalizacion = True
            for caso in valor:
                if x in caso:
                    card = caso.split(x)[0]
                    if card in nvps:
                        nvps.remove(card)
                    if card in para_encontrar_lag:
                        escribir = "MENSAJE: El lag que se encuentra dentro de la VPLS de señalizacion es: " + card.split(':')[0] + " verifique que este es el UPLINK del equipo\n"
                        resultado.write(escribir)
    if nvps:
        print("ADVERTENCIA: La(s) tarjeta(s) NVPS:", nvps, "está(n) fuera de la VPLS", vlan_senalizacion,"de señalizacion")
    if not existe_vpls_senalizacion:
        print("ADVERTENCIA: La VPLS de señalizacion no esta creada o se encuentra shutdown")

def encontrar_lag(datos):
    para_encontrar_lag = []
    for clave in datos.keys():
        var1, var2 = clave.split()[0], clave.split()[1]
        var3 = var1 + '-' +var2 + ':'
        para_encontrar_lag.append(var3)
    return para_encontrar_lag

def verifica_vpls_rtp_incluya_NPOT_LAG_uplink_y_substendido(tarjetas, puertos, lag, servicios, vlan_rtp, sustendido, es_sustendido):
    cantidad_de_lag_en_vpls_rtp = 0
    npot_b = []
    para_encontrar_lag = encontrar_lag(lag)
    existe_vpls_rtp = False
    for clave, valor in tarjetas.items():
        if valor == 'npot-b':
            npot_b.append(clave + ":")
    for clave, valor in servicios.items():
        x = clave.split(" customer")[0].split("vpls ")[-1].strip()
        if vlan_rtp == x:
            existe_vpls_rtp = True
            for caso in valor:
                if x in caso:
                    card = caso.split(x)[0]
                    if card in npot_b:
                        npot_b.remove(card)
                    if card in para_encontrar_lag:
                        cantidad_de_lag_en_vpls_rtp += 1
                        if not es_sustendido:
                            escribir = "MENSAJE:" + card.split(':')[0] + " se encuentra permitido dentro de la VPLS de RTP, verifique si cumple funcion de UPLINK o Subtendido\n"
                            resultado.write(escribir)
                        else:
                            escribir = "MENSAJE:" + card.split(':')[0] + " se encuentra permitido dentro de la VPLS de RTP, verifique que su conexión sea hacia el ISAM principal\n"
                            resultado.write(escribir)
    if not existe_vpls_rtp:
        print("ADVERTENCIA: La VPLS de RTP no esta creada o se encuentra shutdown")
    if npot_b:
        print("ADVERTENCIA: La(s) tarjeta(s) NPOT-B:", npot_b, "está(n) fuera de la VPLS", vlan_rtp,"de RTP")
    if sustendido and cantidad_de_lag_en_vpls_rtp < 2 and not es_sustendido:
        print("ADVERTENCIA: Existe un ISAM Sustendido y no se configuraron 2 LAG en la VPLS de RTP")

def verifica_vprn80008(servicios, senalizacion, rtp):
    retorno_de_funcion_vprn = {}
    virtual_port_senalizacion = 'nt:vp:1:' + senalizacion
    virtual_port_rtp = 'nt:vp:1:' + rtp
    existe_vprn80008 = False
    copia_de_parametros_de_la_vprn = []
    for clave, valor in servicios.items():
        if "vprn 80008" in clave:
            existe_vprn80008 = True
            for caso in valor:
                copia_de_parametros_de_la_vprn.append(caso)
    descripciones = 0
    if len(copia_de_parametros_de_la_vprn) < 11:
        descripciones = 0
        print("ADVERTENCIA: Faltan parametros por configurar en la VPRN, vealos a continuacion:")
        print(copia_de_parametros_de_la_vprn)
        if 'route-distinguisher 1:8' not in copia_de_parametros_de_la_vprn[0]:
            print("Falta parametro 'route-distinguisher 1:8'")
        else:
            for caracteristica in copia_de_parametros_de_la_vprn:
                if '"' in caracteristica:
                    descripciones += 1
            if descripciones < 2:
                if '"' not in copia_de_parametros_de_la_vprn[1]:
                    print("Falta Descripcion de la Megaco-Vlan")
                else:
                    print("Falta Descripcion de la Voice-Vlan")
    if virtual_port_senalizacion in copia_de_parametros_de_la_vprn:
        indice = copia_de_parametros_de_la_vprn.index(virtual_port_senalizacion)
        if indice == 3:
            retorno_de_funcion_vprn.update({'IP5': copia_de_parametros_de_la_vprn[indice - 1]})
            escribir = 'IP5: ' + copia_de_parametros_de_la_vprn[indice - 1] + '\n'
            resultado.write(escribir)
        elif indice < 3:
            print("Falta Direccion IP para la Megaco-Vlan")
    else:
        print("Falta Virtual port para senalizacion:", virtual_port_senalizacion)
    if virtual_port_rtp in copia_de_parametros_de_la_vprn:
        indice = copia_de_parametros_de_la_vprn.index(virtual_port_rtp)
        if indice == 6:
            retorno_de_funcion_vprn.update({'IP3': copia_de_parametros_de_la_vprn[indice - 1]})
            escribir = "IP3: " + copia_de_parametros_de_la_vprn[indice - 1] + '\n'
            resultado.write(escribir)
        elif indice < 6:
            print("Falta Direccion IP para la Voice-VLAN")
    else:
        print("Falta Virtual port para RTP:", virtual_port_rtp )
    ip_maquina_logica = []
    ip_default_router_mgi = []
    for caso in copia_de_parametros_de_la_vprn:
        if 'static-route 0.0.0.0/0' in caso:
            if 'next-hop' in caso:
                retorno_de_funcion_vprn.update({'Default Router RTP': caso.split('next-hop ')[-1].strip()})
                escribir = 'Defaul Router RTP: ' + caso.split('next-hop ')[-1].strip() + '\n'
                resultado.write(escribir)
            else:
                print('Falta Defaul Router de la RTP')
        elif 'static-route' in caso:
            if 'next-hop' in caso:
                escribir = 'MGI: ' + caso.split(' next-hop')[0].split('static-route ')[-1].split('/32')[0] + '\n'
                resultado.write(escribir)
                ip_maquina_logica.append(caso.split(' next-hop')[0].split('static-route ')[-1].split('/32')[0])
                ip_default_router_mgi.append(caso.split('next-hop')[-1])
            else:
                print('Falta Proximo salto para alcanzar la MGI con la IP:', caso.split('static-route ')[-1].strip())
    if len(ip_default_router_mgi)==2:
        if ip_default_router_mgi[0] != ip_default_router_mgi[1]:
            print("ADVERTENCIA: Se configuraron diferentes Default Router para alcanzar las MGI:", ip_default_router_mgi)
        else:
            escribir = "Default Router para alcanzar las MGI con SIG H248: " + ip_default_router_mgi[0] + '\n'
            resultado.write(escribir)
            retorno_de_funcion_vprn.update({'Default Router MGI': ip_default_router_mgi[0]})
    if len(ip_maquina_logica) == 0:
        print("ADVERTENCIA: No se configuraron rutas para acceder a las MGI con la señalización")
    elif len(ip_maquina_logica) == 1:
        print("ADVERTENCIA: No se configuró la ruta para llegar a una de las MGI")
    elif len(ip_maquina_logica) == 2:
        if ip_maquina_logica[0] == ip_maquina_logica[1]:
            print("ADVERTENCIA: Configuró ambas rutas de señalización de la voz hacia la misma MGI")
        else:
            retorno_de_funcion_vprn.update({'MGI A':ip_maquina_logica[0]})
            retorno_de_funcion_vprn.update({'MGI B': ip_maquina_logica[1]})
    else:
        print("ADVERTENCIA: Configuró rutas hacia más de 2 MGI")
    if not existe_vprn80008:
        print("ADVERTENCIA: LA VPRN 80008 no existe o se encuentra shutdown")
    return retorno_de_funcion_vprn

def verifica_creacion_del_media_gateway(media_gateway, vlan_h248, ip_vprn):
    ip2 = []
    if 'vlan-id' in media_gateway:
        vlan = media_gateway.split('vlan-id ')[-1].split(' ')[0].strip()
        if vlan not in vlan_h248:
            print("ADVERTENCIA: La vlan-id que se configuró en el media-gateway no coincide con la vlan de señalización del equipo")
    else:
        print('ADVERTENCIA: No se configuró vlan-id dentro del media-gateway')
    if 'prim-mgc-ip' in media_gateway:
        prim_mgc_ip = media_gateway.split('prim-mgc-ip ')[-1].split()[0].strip()
        if ip_vprn.get('MGI A') != 'NANT-A':
            primaria_en_vprn = ip_vprn.get('MGI A')
            if str(primaria_en_vprn) not in str(prim_mgc_ip):
                secundaria_en_vprn = ip_vprn.get('MGI B')
                if str(secundaria_en_vprn) not in (prim_mgc_ip):
                    print("ADVERTENCIA: No coinciden las MGI primarias configuradas en la VPRN 80008 y el media-gateway, vea los detalles: VPRN 80008:", primaria_en_vprn, secundaria_en_vprn, "Media-gateawy:", prim_mgc_ip)
    else:
        print("ADVERTENCIA: No se configuró prim-mgc-ip en el media-gateway")
    if "name" in media_gateway:
        name = media_gateway.split(' name ')[-1].split()[0].strip()
        escribir = "Nombre del equipo en el media-gateway debe coincidir con el MGC: " + name + '\n'
        ip2.append(name)
        resultado.write(escribir)
    else:
        print("ADVERTENCIA: No se configuró el nombre del equipo en el media-gateway, tenga en cuenta que debe coincidir con el nombre del equipo en el MGC")
    if "ip-address" in media_gateway:
        ip_address = media_gateway.split("ip-address")[-1].split()[0]
        if "netmask" in media_gateway:
            mascara = media_gateway.split('netmask ')[-1]. split()[0].strip()
            ip2.append(ip_address)
            ip2.append(mascara)
            escribir = "IP2: " + ip_address + " máscara: " + mascara + '\n'
            resultado.write(escribir)
        else:
            escribir = "IMPORTANTE, verifique que esta IP coincida con la IP2: " + ip_address + '\n'
            resultado.write(escribir)
            print("ADVERTENCIA: No se configuró máscara para la IP2")
    else:
        print("ADVERTENCIA: No se configuró la IP2 en el media-gateway")
    if 'router-ip' in media_gateway:
        router_ip = media_gateway.split("router-ip ")[-1]. split()[0].strip()
        if ip_vprn.get('Default Router MGI') != 'NANT-A':
            default_router_mgi_en_vprn = ip_vprn.get('Default Router MGI')
            if default_router_mgi_en_vprn is not None:
                if router_ip not in default_router_mgi_en_vprn:
                    print("ADVERTENCIA: La IP del router-ip configurada en el media-gateway:", router_ip ,"no coincide con el Default Router MGI configurado en la VPRN:", default_router_mgi_en_vprn)
    else:
        print("ADVERTENCIA: No se configuró router-ip en el media-gateway, este es el Default Router para acceder a las MGI")
    if 'sec-mgc-ip' in media_gateway:
        sec_mgc_ip = media_gateway.split("sec-mgc-ip ")[-1].split()[0]
        if ip_vprn.get("MGI B") != 'NANT-A':
            mgi_b_vprn = ip_vprn.get("MGI B")
            if mgi_b_vprn is not None:
                if sec_mgc_ip not in mgi_b_vprn:
                    secundaria_en_vprn = ip_vprn.get('MGI A')
                    if sec_mgc_ip not in secundaria_en_vprn:
                        print("ADVERTENCIA: No coinciden las MGI secundarias configuradas en la VPRN 80008 y el media-gateway, vea los detalles: VPRN 80008:",mgi_b_vprn, secundaria_en_vprn, "Media-gateawy:", sec_mgc_ip)
    else:
        print("ADVERTENCIA: No se configuró una MGI secundaria")
    if 'admin-status unlocked' not in media_gateway:
        print("ADVERTENCIA: No se ha desbloqueado el media-gateway por lo que no está operativo el servicio de voz en el ISAM")
    if 'isdn-term-format BA isdn-suffix1 /B1 isdn-suffix2 /B2' not in media_gateway:
        print("ADVERTENCIA: No se configuraron los siguientes parámetros del media-gateway: isdn-term-format BA isdn-suffix1 /B1 isdn-suffix2 /B2")
    return ip2

def verifica_creacion_del_cluster(cluster, rtp, vprn):
    ip4 = []
    ip = cluster.split('ivps-ip ')[-1].split()[0]
    if 'netmask' in cluster:
        mascara = cluster.split('netmask ')[-1].split()[0]
    else:
        print("ADVERTENCIA: En la creacion del cluster no se especificó máscara para la IP4")
    if 'router-ip' in cluster:
        router = cluster.split('router-ip ')[-1].split()[0].strip()
        if vprn.get('Default Router RTP') != 'NANT-A':
            router_en_vprn = vprn.get('Default Router RTP')
            if router not in router_en_vprn:
                print("ADVERTENCIA: No coinciden los Default Router para RTP entre el cluster:", router, "y la VPRN:", router_en_vprn)
    else:
        print("ADVERTENCIA: No se ha configurado Default Router para RTP al crear el cluster de voz")
    if "vlan-id" in cluster:
        vlan = cluster.split('vlan-id ')[-1].strip()
        if vlan not in rtp:
            print('ADVERTENCIA: La Vlan de RTP configurada en el cluster no coincide con la establecida para el equipo')
    else:
        print("ADVERTENCIA: No se configuró Vlan de RTP en la creación del cluster de voz")
    ip4.append(ip)
    ip4.append(mascara)
    escribir = 'IP4: ' + ip + ' máscara: ' + mascara + '\n'
    resultado.write(escribir)
    return ip4

def verifica_adicion_de_equipment_al_cluster(mando, name, vprn):
    respuesta_de_equipment_en_cluster = []
    for caso in mando:
        if 'equipment 1' in caso:
            asam_id = caso.split('asam-id ')[-1].split()[0].strip()
            if asam_id != name:
                print("ADVERTENCIA: El nombre del equipment que se configuró en el cluster no coincide con el System Name del ISAM Principal:",asam_id, name)
            if 'ip-address' in caso and 'IP3' in vprn.keys():
                ip_3 = caso.split('ip-address ')[-1].strip()
                if ip_3 != vprn.get('IP3').split('/24')[0]:
                    print("ADVERTENCIA: No coincide la IP3 configurada al adicionar el equipment 1 al cluster:", ip_3, "con la IP3 configurada en la VPRN:", vprn.get('IP3').split('/24')[0])
            else:
                print("ADVERTENCIA: No se configuró la IP3 del ISAM principal en el cluster")
        if 'equipment 2' in caso:
            asam_id = caso.split('asam-id ')[-1].split()[0].strip()
            respuesta_de_equipment_en_cluster.append(asam_id)
            if 'ip-address' in caso:
                ip_3 = caso.split('ip-address ')[-1].strip()
                respuesta_de_equipment_en_cluster.append(ip_3)
            else:
                print("ADVERTENCIA: No se configuró la IP3 del ISAM sustendido al adicionarlo al cluster")
    return respuesta_de_equipment_en_cluster

def verifica_tarjetas_NPOTB_se_incluyan_en_cluster(ranuras, cluster, equipment):
    x = cluster.get(equipment)
    for clave, valor in ranuras.items():
        if 'npot-b' in valor:
            if clave.split('lt:')[-1] in x:
                x.remove(clave.split('lt:')[-1])
            else:
                print("ADVERTENCIA: La tarjeta NPOT-B ubicada en la ranura:", clave, "no fue adicionada al cluster de voz")
    if x:
        print("ADVERTENCIA: La(s) siguientes(s) tarjeta(s) se encuentra(n) dentro del cluster y no son NPOT-B")

def verifica_terminaciones_en_cluster_de_voz(termination, cluster, equipo):
    x = cluster.get(equipo)
    y = termination.get(equipo)
    resultado.write("IMPORTANTE, verifique la correspondencia entre el puerto y la terminación con los datos siguientes:\n")
    orden = -1
    if x:
        for tarjeta in x:
            for numero in range(1,73):
                orden += 1
                respuesta = False
                term = tarjeta + '/' + str(numero)
                if numero >= 9:
                    respuesta = chequea_terminaciones_que_deben_exitir_vs_creadas(term, y, False)
                else:
                    respuesta = chequea_terminaciones_que_deben_exitir_vs_creadas(term, y, True)
                if respuesta:
                    if orden == 0:
                        consecutivo = int(respuesta) + orden
                        escribir = term + ',' + respuesta + '\n'
                        resultado.write(escribir)
                    else:
                        if consecutivo + orden == int(respuesta):
                            escribir = term + ',' + respuesta + '\n'
                            resultado.write(escribir)
                        else:
                            consecutivo = int(respuesta)
                            orden = 0
                            escribir = "CUIDADO SALTO EN EL ORDEN" + '\n'
                            resultado.write(escribir)
                            escribir = term + ',' + respuesta + '\n'
                            resultado.write(escribir)

def chequea_terminaciones_que_deben_exitir_vs_creadas(terminacion, term_creadas, chequear_telef_public):
    valor_a_retornar = False
    for caso in term_creadas:
        if terminacion in caso:
            if 'admin-status unlocked' not in caso:
                print("ADVERTENCIA: La terminación:", terminacion, "no se encuentra desbloqueada")
            if chequear_telef_public:
                if 'line-feed 40' not in caso:
                    print("ADVERTENCIA: La terminación:", terminacion, "no fue configurada para teléfono público")
            if 'termination-id' not in caso:
                print("ADVERTENCIA: no fue especificada la terminación para", terminacion)
            else:
                valor_a_retornar = caso.split('termination-id ')[-1].split()[0].strip()
            break
    if not valor_a_retornar:
        print("ADVERTENCIA: No fue creada la terminación del slot:", terminacion)
    return valor_a_retornar

def verifica_vprn80008_en_sustendido(servicios, rtp):
    retorno_de_funcion_vprn = {}
    virtual_port_rtp = 'nt:vp:1:' + rtp
    existe_vprn80008 = False
    copia_de_parametros_de_la_vprn = []
    for clave, valor in servicios.items():
        if "vprn 80008" in clave:
            existe_vprn80008 = True
            for caso in valor:
                copia_de_parametros_de_la_vprn.append(caso)
    descripciones = 0
    if len(copia_de_parametros_de_la_vprn) != 6:
        descripciones = 0
        print("ADVERTENCIA: Faltan parametros por configurar en la VPRN, vealos a continuacion:")
        if 'route-distinguisher 1:8' not in copia_de_parametros_de_la_vprn[0]:
            print("Falta parametro 'route-distinguisher 1:8'")
    for caracteristica in copia_de_parametros_de_la_vprn:
        if '"' in caracteristica:
            descripciones += 1
    if descripciones < 1:
        print("Falta Descripcion de la Voice-Vlan")
    if virtual_port_rtp in copia_de_parametros_de_la_vprn:
        indice = copia_de_parametros_de_la_vprn.index(virtual_port_rtp)
        if indice == 3:
            retorno_de_funcion_vprn.update({'IP3': copia_de_parametros_de_la_vprn[indice - 1]})
            escribir = "IP3 del ISAM SUSTENDIDO:" + copia_de_parametros_de_la_vprn[indice - 1] + '\n'
            resultado.write(escribir)
        elif indice < 3:
            print("Falta Direccion IP para la Voice-VLAN")
    else:
        print("Falta Virtual port para RTP:", virtual_port_rtp )
    retorno = verifica_ruta_y_salto_en_vprn(copia_de_parametros_de_la_vprn)
    if retorno:
        retorno_de_funcion_vprn.update(retorno)
    return retorno_de_funcion_vprn

def verifica_ruta_y_salto_en_vprn(datos):
    a_devolver = {}
    for caso in datos:
        if 'static-route 0.0.0.0/0' in caso:
            if 'next-hop' in caso:
                a_devolver.update({'Default Router RTP': caso.split('next-hop ')[-1].strip()})
                escribir = 'Defaul Router RTP: ' + caso.split('next-hop ')[-1].strip() + '\n'
                resultado.write(escribir)
            else:
                print('Falta Defaul Router de la RTP')
    return a_devolver

###############################----------------------MAIN-----------------############################################
#INSTRUCCIONES
#Coloque el(los) archivo(s) de configuracion en la misma carpeta del script, si existe 1 ISAM sustendido pongale como nombre "sustendido.txt"
#De existir varios ISAM sustendidos diferencielos usando numeros en el nombre Ejemplo: "sustendido1.txt", "sustendido2.txt" etc
vlan_rtp = "1410"
vlan_senalizacion = "1210"
vlan_gestion = "180"
####################################
resultado = open('Parametros_configurados.txt', 'w')

resultado.write("Los siguientes ficheros fueron leídos de la carpeta: " + os.getcwd() + '\n')
ficheros_en_carpeta = glob.glob('*.txt') + glob.glob('*.log')
if ficheros_en_carpeta:
    for caso in ficheros_en_carpeta:
        if caso != 'Parametros_configurados.txt':
            resultado.write(caso + '\n')
else:
    print('No existen archivos con extension txt o log en la carpeta')

excel_con_datos = glob.glob('*.xlsx')
print(excel_con_datos)
if len(excel_con_datos) > 1:
    print('ADVERTENCIA: No debe existir más de un documento excel con el direccionamiento IP del sitio, rectifique esto y ejecute el script nuevamente')
else:
    file_path = excel_con_datos[0]
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    for i in range(1, max_row +1):
        for j in range(1, max_column + 1):
            if i == 2 and j == max_column:
                if sheet.cell(row = 2, column = j).value is not None and 'endido' in sheet.cell(row = 2, column = j).value:
                    print('EXCEL:',sheet.cell(row = 2, column = j).value)
    wb.close()
    tarjetas = {}
    para_analizar_puertos = []
    para_analizar_LAG = []
    para_analizar_servicios = []
    equipos_en_el_cluster = 0
    existe_sustendido = False
    debe_existir_lag = False
    adicion_de_equipment_al_cluster = []
    tarjetas_en_el_cluster = {}
    cards = []
    tarj = []
    terminaciones_principal = []
    terminaciones_sustendido = []
    terminaciones = {}
    comunidades_snmp = []
    para_analizar_vlan = []
    para_analizar_ip_gestion = []
    para_analizar_vrf = []
    para_analizar_interface = []
    datos_del_sustendido_en_el_cluster = ''
    ip_configuradas_en_vprn = {}
    for nombre in ficheros_en_carpeta:
        if "sustendido" not in nombre:
            if os.path.isfile(nombre):
                # f_resultfile = open("resultado.txt", "w")
                f_logfile = open(nombre, "r", encoding="latin-1").readlines()
                for cuenta_linea, linea in enumerate(f_logfile):
                    #Capta el tipo de tarjeta presente en cada ranura
                    if "configure system id" in linea and "name" in linea:
                        system_id = linea.split('system id ')[-1].split()[0].strip()
                        system_name = linea.split("name ")[-1].split()[0].strip()
                        if system_id != system_name:
                            print("ADVERTENCIA: No coinciden el System Name y el System Id, esto provocará que el equipo no funcione o lo haga MAL")
                    if "configure equipment slot" in linea and "unlock" in linea:
                        if "planned-type" in linea:
                            ranura = linea.split("slot ")[-1].split()[0].strip()
                            tarjeta = linea.split("planned-type ")[-1].split()[0].strip()
                            tarjetas.update({ranura: tarjeta})
                    if 'configure voice cluster 1 media-gateway 1' in linea:
                        configuracion_del_media_gateway = linea.strip()
                    if 'ip ivps-ip' in linea:
                        creacion_del_cluster = linea.strip()
                    if "configure voice cluster 1 equipment" in linea and 'asam-id' in linea:
                        equipos_en_el_cluster += 1
                        adicion_de_equipment_al_cluster.append(linea.strip())
                    if "configure voice cluster 1 equipment" in linea and "board" in linea:
                        if 'equipment 1' in linea:
                            cards.append(linea.split('board ')[-1].split()[0].strip())
                        elif 'equipment 2' in linea:
                            tarj.append(linea.split('board ')[-1].split()[0].strip())
                    if "configure voice cluster 1 equipment" in linea and 'termination' in linea:
                        if "equipment 1" in linea:
                            terminaciones_principal.append(linea.split('termination ')[-1].strip())
                        elif "equipment 2" in linea:
                            terminaciones_sustendido.append(linea.split('termination ')[-1].strip())
                    if 'configure system security snmp community' in linea:
                        comunidades_snmp.append(linea.split('configure system security snmp community')[-1].strip())
                for cuenta_linea, linea in enumerate(f_logfile):
                    tipo_de_isam = tarjetas.get('nt-a')
                    if tipo_de_isam == 'nant-e':
                        if 'echo "Port Configuration"' in linea:
                            for rango_siguiente, lineas_siguientes in enumerate(f_logfile[cuenta_linea + 2:]):
                                if "#--------------------------------------------------" in lineas_siguientes:
                                    break
                                else:
                                    para_analizar_puertos.append(lineas_siguientes.strip())
                            puertos = analiza_puertos(para_analizar_puertos)
                        if 'echo "LAG Configuration"' in linea:
                            for rango_siguiente, lineas_siguientes in enumerate(f_logfile[cuenta_linea + 2:]):
                                if "#--------------------------------------------------" in lineas_siguientes:
                                    break
                                else:
                                    para_analizar_LAG.append(lineas_siguientes.strip())
                            lag = analiza_lag(para_analizar_LAG)
                        if 'echo "Service Configuration"' in linea:
                            for rango_siguiente, lineas_siguientes in enumerate(f_logfile[cuenta_linea + 2:]):
                                if "#--------------------------------------------------" in lineas_siguientes:
                                    break
                                else:
                                    para_analizar_servicios.append(lineas_siguientes.strip())
                            servicios = analiza_servicios(para_analizar_servicios, vlan_rtp, vlan_senalizacion, tarjetas, False)
                    elif tipo_de_isam == 'nant-a':
                        if 'echo "vlan"' in linea:
                            for rango_siguiente, lineas_siguientes in enumerate(f_logfile[cuenta_linea + 2:]):
                                if "#----------------------------------------------------------------------------" in lineas_siguientes:
                                    break
                                else:
                                    para_analizar_vlan.append(lineas_siguientes.strip())
                        if 'echo "system"' in linea:
                            for rango_siguiente, lineas_siguientes in enumerate(f_logfile[cuenta_linea + 2:]):
                                if "#----------------------------------------------------------------------------" in lineas_siguientes:
                                    break
                                else:
                                    para_analizar_ip_gestion.append(lineas_siguientes.strip())
                        if 'echo "ip"' in linea:
                            for rango_siguiente, lineas_siguientes in enumerate(f_logfile[cuenta_linea + 2:]):
                                if "#----------------------------------------------------------------------------" in lineas_siguientes:
                                    break
                                else:
                                    para_analizar_vrf.append(lineas_siguientes.strip())
                        if 'echo "interface"' in linea:
                            for rango_siguiente, lineas_siguientes in enumerate(f_logfile[cuenta_linea + 2:]):
                                if "#----------------------------------------------------------------------------" in lineas_siguientes:
                                    break
                                else:
                                    para_analizar_interface.append(lineas_siguientes.strip())
        else:
            existe_sustendido = True

    resultado.write("**********************************  En el ISAM Cabeza de grupo: ******************************\n")
    if terminaciones_principal:
        terminaciones.update({'equipment 1': terminaciones_principal})
    else:
        if cards:
            print('ADVERTENCIA: No se han adicionado las terminaciones del ISAM principal al cluster de voz')
    if terminaciones_sustendido:
        terminaciones.update({'equipment 2': terminaciones_sustendido})
    if cards:
        tarjetas_en_el_cluster.update({'equipment 1': cards})
    else:
        if cards:
            print("ADVERTENCIA: Las tarjetas del equipment 1 no han sido añadidas al cluster de voz")
    if tarj:
        tarjetas_en_el_cluster.update({'equipment 2': tarj})
    else:
        print("ADVERTENCIA: Las tarjetas del equipment 2 no han sido añadidas al cluster de voz")
    if 'nant-e' in tarjetas.get('nt-a'):
        if len(lag) == 0:
            print("ADVERTENCIA: No existe LAG configurado en el equipo")
        elif len(lag) == 1:
            if existe_sustendido:
                print("ADVERTENCIA: Existe un ISAM Sustendido y solo se ha configurado un LAG en el equipo")
        ip_gestion_principal = verifica_gestion_del_equipo(servicios, vlan_gestion, tarjetas, lag)
        verifica_puertos_del_lag_operativos(puertos, lag)
        verifica_operativos_puertos_de_tarjetas_y_virtual_port(tarjetas, puertos)
        verifica_vpls_senalizacion_incluya_NVPS_LAG_puerto_uplink(tarjetas, puertos, lag, servicios, vlan_senalizacion)
        verifica_vpls_rtp_incluya_NPOT_LAG_uplink_y_substendido(tarjetas, puertos, lag, servicios, vlan_rtp, existe_sustendido, False)
        ip_configuradas_en_vprn = verifica_vprn80008(servicios, vlan_senalizacion, vlan_rtp)
        if configuracion_del_media_gateway:
            ip2 = verifica_creacion_del_media_gateway(configuracion_del_media_gateway, vlan_senalizacion, ip_configuradas_en_vprn)
        else:
            print('ADVERTENCIA: No se ha configurado el Media Gateway, requisito imprescindible en la configuración de voz')
        if creacion_del_cluster:
            ip4 = verifica_creacion_del_cluster(creacion_del_cluster, vlan_rtp, ip_configuradas_en_vprn)
        else:
            print("ADVERTENCIA: No se ha creado el cluster de voz en el ISAM")
        if equipos_en_el_cluster == 0:
            print("ADVERTENCIA:No se han adicionado equipos al cluster de voz")
        else:
            if equipos_en_el_cluster == 1 and existe_sustendido:
                print("ADVERTENCIA: Existe un ISAM sustendido y solo se ha adicionado un equipo al cluster de voz")
            else:
                datos_del_sustendido_en_el_cluster = verifica_adicion_de_equipment_al_cluster(adicion_de_equipment_al_cluster,system_name, ip_configuradas_en_vprn)
    if 'nant-a' in tarjetas.get('nt-a'):
        verifica_puerto_y_tarjetas_dentro_de_las_vlan_senalizacion_rtp_gestion_nant_a(para_analizar_vlan, vlan_gestion, vlan_rtp, vlan_senalizacion, tarjetas, existe_sustendido)
        ip_gestion_nant_a = devuelve_ip_gestion_en_nanta(para_analizar_ip_gestion)
        if ip_gestion_nant_a:
            ip_gestion_principal = ip_gestion_nant_a.get('ip1')
            mascara_ip_gestion_principal = ip_gestion_nant_a.get('Default route gestion')
        default_router_rtp = analiza_vrf(para_analizar_vrf, vlan_rtp)
        ip3 = analiza_interface_en_nanta(para_analizar_interface, vlan_rtp)
        ip_configuradas_en_vprn.update({'IP5':'NANT-A'})
        ip_configuradas_en_vprn.update({'Default Router MGI': 'NANT-A'})
        ip_configuradas_en_vprn.update({'MGI A':'NANT-A'})
        ip_configuradas_en_vprn.update({'MGI B': 'NANT-A'})
        ip_configuradas_en_vprn.update(default_router_rtp)
        ip_configuradas_en_vprn.update(ip3)
        if configuracion_del_media_gateway:
            ip2 = verifica_creacion_del_media_gateway(configuracion_del_media_gateway, vlan_senalizacion, ip_configuradas_en_vprn)
        else:
            print('ADVERTENCIA: No se ha configurado el Media Gateway, requisito imprescindible en la configuración de voz')
        if creacion_del_cluster:
            ip4 = verifica_creacion_del_cluster(creacion_del_cluster, vlan_rtp, ip_configuradas_en_vprn)
        else:
            print("ADVERTENCIA: No se ha creado el cluster de voz en el ISAM")
        if equipos_en_el_cluster == 0:
            print("ADVERTENCIA:No se han adicionado equipos al cluster de voz")
        else:
            if equipos_en_el_cluster == 1 and existe_sustendido:
                print("ADVERTENCIA: Existe un ISAM sustendido y solo se ha adicionado un equipo al cluster de voz")
            else:
                datos_del_sustendido_en_el_cluster = verifica_adicion_de_equipment_al_cluster(adicion_de_equipment_al_cluster,system_name, ip_configuradas_en_vprn)
    copia_de_tarjetas_en_el_cluster = copy.deepcopy(tarjetas_en_el_cluster)
    verifica_tarjetas_NPOTB_se_incluyan_en_cluster(tarjetas, copia_de_tarjetas_en_el_cluster, 'equipment 1')
    copia_de_tarjetas_en_el_cluster = copy.deepcopy(tarjetas_en_el_cluster)
    copia_de_terminaciones_en_el_cluster = copy.deepcopy(terminaciones)
    verifica_terminaciones_en_cluster_de_voz(terminaciones, tarjetas_en_el_cluster, "equipment 1")


    para_analizar_puertos_sustendido = []
    para_analizar_LAG_sustendido = []
    para_analizar_servicios_sustendido = []
    tarjetas_sustendido = {}
    comunidades_snmp_sustendido = []
    para_analizar_vlan_sustendido = []
    para_analizar_ip_gestion_sustendido = []
    para_analizar_vrf_sustendido = []
    para_analizar_interface_sustendido = []
    for nombre in ficheros_en_carpeta:
        if "sustendido" in nombre:
            existe_sustendido = True
            if os.path.isfile(nombre):
                # f_resultfile = open("resultado.txt", "w")
                f_logfile = open(nombre, "r", encoding="latin-1").readlines()
                for cuenta_linea, linea in enumerate(f_logfile):
                    #Capta el tipo de tarjeta presente en cada ranura
                    if "configure system id" in linea and "name" in linea:
                        system_id = linea.split('system id ')[-1].split()[0].strip()
                        system_name = linea.split("name ")[-1].split()[0].strip()
                        if system_id != system_name:
                            print("ADVERTENCIA: No coinciden el System Name y el System Id, esto provocará que el equipo no funcione o lo haga MAL")
                    if "configure equipment slot" in linea and "unlock" in linea:
                        if "planned-type" in linea:
                            ranura_sustendido = linea.split("slot ")[-1].split()[0].strip()
                            tarjeta_sustendido = linea.split("planned-type ")[-1].split()[0].strip()
                            tarjetas_sustendido.update({ranura_sustendido: tarjeta_sustendido})
                    if 'configure system security snmp community' in linea:
                        comunidades_snmp_sustendido.append(linea.split('configure system security snmp community')[-1].strip())
                tipo_de_isam = tarjetas_sustendido.get('nt-a')
                for cuenta_linea, linea in enumerate(f_logfile):
                    if tipo_de_isam == 'nant-e':
                        if 'echo "Port Configuration"' in linea:
                            for rango_siguiente, lineas_siguientes in enumerate(f_logfile[cuenta_linea + 2:]):
                                if "#--------------------------------------------------" in lineas_siguientes:
                                    break
                                else:
                                    para_analizar_puertos_sustendido.append(lineas_siguientes.strip())
                            puertos_sustendido = analiza_puertos(para_analizar_puertos_sustendido)
                        if 'echo "LAG Configuration"' in linea:
                            for rango_siguiente, lineas_siguientes in enumerate(f_logfile[cuenta_linea + 2:]):
                                if "#--------------------------------------------------" in lineas_siguientes:
                                    break
                                else:
                                    para_analizar_LAG_sustendido.append(lineas_siguientes.strip())
                            lag_sustendido = analiza_lag(para_analizar_LAG_sustendido)
                        if 'echo "Service Configuration"' in linea:
                            for rango_siguiente, lineas_siguientes in enumerate(f_logfile[cuenta_linea + 2:]):
                                if "#--------------------------------------------------" in lineas_siguientes:
                                    break
                                else:
                                    para_analizar_servicios_sustendido.append(lineas_siguientes.strip())
                            servicios_sustendido = analiza_servicios(para_analizar_servicios_sustendido, vlan_rtp, vlan_senalizacion, tarjetas_sustendido, True)
                    elif tipo_de_isam == 'nant-a':
                        if 'echo "vlan"' in linea:
                            for rango_siguiente, lineas_siguientes in enumerate(f_logfile[cuenta_linea + 2:]):
                                if "#---------------------------------------------------------------------------------------------------" in lineas_siguientes:
                                    break
                                else:
                                    para_analizar_vlan_sustendido.append(lineas_siguientes.strip())

                        if 'echo "system"' in linea:
                            for rango_siguiente, lineas_siguientes in enumerate(f_logfile[cuenta_linea + 2:]):
                                if "#---------------------------------------------------------------------------------------------------" in lineas_siguientes:
                                    break
                                else:
                                    para_analizar_ip_gestion_sustendido.append(lineas_siguientes.strip())
                        if 'echo "ip"' in linea:
                            for rango_siguiente, lineas_siguientes in enumerate(f_logfile[cuenta_linea + 2:]):
                                if "#---------------------------------------------------------------------------------------------------" in lineas_siguientes:
                                    break
                                else:
                                    para_analizar_vrf_sustendido.append(lineas_siguientes.strip())
                        if 'echo "interface"' in linea:
                            for rango_siguiente, lineas_siguientes in enumerate(f_logfile[cuenta_linea + 2:]):
                                if "#---------------------------------------------------------------------------------------------------" in lineas_siguientes:
                                    break
                                else:
                                    para_analizar_interface_sustendido.append(lineas_siguientes.strip())
    if existe_sustendido:
        resultado.write("\n**********************************  En el ISAM Sustendido: ******************************\n")
        if tipo_de_isam == 'nant-e':
            ip_gestion_sustendido = verifica_gestion_del_equipo(servicios_sustendido, vlan_gestion, tarjetas_sustendido, lag_sustendido)
            if puertos_sustendido and lag_sustendido:
                verifica_puertos_del_lag_operativos(puertos_sustendido, lag_sustendido)
            if tarjetas_sustendido and puertos_sustendido:
                verifica_operativos_puertos_de_tarjetas_y_virtual_port(tarjetas_sustendido, puertos_sustendido)
            if tarjetas_sustendido and puertos_sustendido and lag_sustendido and servicios_sustendido and existe_sustendido:
                verifica_vpls_rtp_incluya_NPOT_LAG_uplink_y_substendido(tarjetas_sustendido, puertos_sustendido, lag_sustendido, servicios_sustendido, vlan_rtp, existe_sustendido, True)
            copia_de_tarjetas_en_el_cluster = copy.deepcopy(tarjetas_en_el_cluster)
            verifica_tarjetas_NPOTB_se_incluyan_en_cluster(tarjetas_sustendido, copia_de_tarjetas_en_el_cluster, 'equipment 2')
            if datos_del_sustendido_en_el_cluster:
                if datos_del_sustendido_en_el_cluster[0] != system_name:
                    print("ADVERTENCIA: El asam-id que se configuró para el equipment 2", datos_del_sustendido_en_el_cluster[0] ,"no coincide con el System Name del SUSTENDIDO:", system_name)
            ip_configuradas_en_vprn_sustendido = verifica_vprn80008_en_sustendido(servicios_sustendido, vlan_rtp)
            ip3 = ip_configuradas_en_vprn_sustendido.get('IP3')
            if datos_del_sustendido_en_el_cluster:
                if datos_del_sustendido_en_el_cluster[1] != ip3.split('/24')[0].strip():
                    print("ADVERTENCIA: La IP3 que se configuró al adicionar el equipment 2 al cluster de voz", datos_del_sustendido_en_el_cluster[1],"no coincide con la IP3 configurada en la VPRN 80008 del SUSTENDIDO:", ip3.split('/24')[0].strip())
            verifica_terminaciones_en_cluster_de_voz(terminaciones, tarjetas_en_el_cluster, "equipment 2")
        elif tipo_de_isam == 'nant-a':
            verifica_puerto_y_tarjetas_dentro_de_las_vlan_senalizacion_rtp_gestion_nant_a(para_analizar_vlan_sustendido, vlan_gestion, vlan_rtp, vlan_senalizacion, tarjetas_sustendido, existe_sustendido)
            ip_gestion_nant_a = devuelve_ip_gestion_en_nanta(para_analizar_ip_gestion_sustendido)
            if ip_gestion_nant_a:
                ip_gestion_principal = ip_gestion_nant_a.get('ip1')
                mascara_ip_gestion_principal = ip_gestion_nant_a.get('Default route gestion')
            default_router_rtp = analiza_vrf(para_analizar_vrf_sustendido, vlan_rtp)
            ip3 = analiza_interface_en_nanta(para_analizar_interface_sustendido, vlan_rtp)
            ip_configuradas_en_vprn.update({'IP5': 'NANT-A'})
            ip_configuradas_en_vprn.update({'Default Router MGI': 'NANT-A'})
            ip_configuradas_en_vprn.update({'MGI A': 'NANT-A'})
            ip_configuradas_en_vprn.update({'MGI B': 'NANT-A'})
            ip_configuradas_en_vprn.update(default_router_rtp)
            ip_configuradas_en_vprn.update(ip3)
            if configuracion_del_media_gateway:
                ip2 = verifica_creacion_del_media_gateway(configuracion_del_media_gateway, vlan_senalizacion,ip_configuradas_en_vprn)
            else:
                print(
                    'ADVERTENCIA: No se ha configurado el Media Gateway, requisito imprescindible en la configuración de voz')
            if creacion_del_cluster:
                ip4 = verifica_creacion_del_cluster(creacion_del_cluster, vlan_rtp, ip_configuradas_en_vprn)
            else:
                print("ADVERTENCIA: No se ha creado el cluster de voz en el ISAM")
            if equipos_en_el_cluster == 0:
                print("ADVERTENCIA:No se han adicionado equipos al cluster de voz")
            else:
                if equipos_en_el_cluster == 1 and existe_sustendido:
                    print("ADVERTENCIA: Existe un ISAM sustendido y solo se ha adicionado un equipo al cluster de voz")
                else:
                    datos_del_sustendido_en_el_cluster = verifica_adicion_de_equipment_al_cluster(adicion_de_equipment_al_cluster, system_name, ip_configuradas_en_vprn)
            copia_de_tarjetas_en_el_cluster = copy.deepcopy(tarjetas_en_el_cluster)
            verifica_tarjetas_NPOTB_se_incluyan_en_cluster(tarjetas_sustendido, copia_de_tarjetas_en_el_cluster,'equipment 2')
            verifica_terminaciones_en_cluster_de_voz(terminaciones, tarjetas_en_el_cluster, "equipment 2")
    resultado.close()

end_time = time()
elapsed = end_time - start_time
print(elapsed)