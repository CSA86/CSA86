import openpyxl
import re

file_path = 'Listado_de_puertas.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active
max_row = sheet.max_row
max_column = sheet.max_column
sheet.cell(row=1, column=max_column + 1).value = 'Tipo de Puerta'
sheet.cell(row=1, column=max_column + 2).value = 'Tipo de conectividad'

wb.save(file_path)
c = 0
for i in range(1, max_column + 1):
    nombre_del_campo = sheet.cell(row=1, column=i).value
    if nombre_del_campo == 'PUERTA':
        indice_de_columna_puerta = i
    if nombre_del_campo == 'TIPO':
        indice_de_columna_tipo = i
    if nombre_del_campo == 'SERVICIO':
        indice_de_columna_servicio = i
    if nombre_del_campo == 'HOGAR':
        indice_de_columna_hogar = i
    if nombre_del_campo == 'ESTADO':
        indice_de_columna_estado = i

for j in range(2,max_row + 1):
    valor_del_campo = sheet.cell(row=j, column=indice_de_columna_tipo).value
    valor_a_dividir = sheet.cell(row=j, column=indice_de_columna_puerta).value
    if valor_del_campo == 'DSLAM IP':
        temporal = re.findall('ADSL|VDSL|SHDSL|ETH|GETH|GE/OP|GE/FE', valor_a_dividir)
        longitud = len(temporal)
        if longitud == 2:
            tipo_de_puerta = temporal[1]
            tipo_de_puerta = tipo_de_puerta.replace('/', '')
            if tipo_de_puerta == 'ADSL' or tipo_de_puerta == 'VDSL' or tipo_de_puerta == 'SHDSL':
                sheet.cell(row=j, column=max_column + 1).value = tipo_de_puerta
            else:
                sheet.cell(row=j, column=max_column + 1).value = 'OTRAS'
        else:
            if temporal:
                tipo_de_puerta = temporal[0]
                if 'GE/OP' in tipo_de_puerta or 'GE/FE' in tipo_de_puerta:
                    pass
                else:
                    tipo_de_puerta = tipo_de_puerta.replace('/', '')
                if tipo_de_puerta == 'ADSL' or tipo_de_puerta == 'VDSL' or tipo_de_puerta == 'SHDSL':
                    sheet.cell(row=j, column=max_column + 1).value = tipo_de_puerta
                else:
                    sheet.cell(row=j, column=max_column + 1).value = 'OTRAS'
            else:
                sheet.cell(row=j, column=max_column + 1).value = 'OTRAS'
    else:
        sheet.cell(row=j, column=max_column + 1).value = 'OTRAS'
        if valor_del_campo == 'CN4':
            sheet.cell(row=j, column=indice_de_columna_estado).value = 'OCUPADA'
        elif valor_del_campo == 'E1 CE':
            sheet.cell(row=j, column=indice_de_columna_estado).value = 'OCUPADA'
        elif valor_del_campo == 'E1 CFR':
            sheet.cell(row=j, column=indice_de_columna_estado).value = 'OCUPADA'
        elif valor_del_campo == 'E1 UFR':
            sheet.cell(row=j, column=indice_de_columna_estado).value = 'OCUPADA'
        elif valor_del_campo == 'HDE1FR':
            sheet.cell(row=j, column=indice_de_columna_estado).value = 'OCUPADA'
        elif valor_del_campo == 'STM1 IR':
            sheet.cell(row=j, column=indice_de_columna_estado).value = 'OCUPADA'
for k in range(2, max_row + 1):
    valor = sheet.cell(row=k, column=indice_de_columna_hogar).value
    if valor == 'NO':
        sheet.cell(row=k, column=max_column + 2).value = 'Conectividad Social'
    elif valor == 'SI':
        sheet.cell(row=k, column=max_column + 2).value = 'NAUTA HOGAR'
wb.save(file_path)

