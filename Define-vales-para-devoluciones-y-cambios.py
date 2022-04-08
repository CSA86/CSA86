import openpyxl
import glob
import itertools
import copy
import timeit
import datetime
from operator import itemgetter, attrgetter

start = timeit.timeit()

def analisis_de_recursos_donde_consumo_excede_salida_almacen(recursos, hoja, titulo_columna):
    maxima_fila = hoja.max_row
    maxima_columna = hoja.max_column
    vales_por_recurso = {}
    for i in range(1, maxima_columna + 1):
        valor = hoja.cell(row=titulo_columna, column=i).value
        if valor == 'CMv':
            columna_movimiento = i
        if valor == 'Doc.mat.':
            columna_vale = i
        if valor == 'Texto breve de material':
            columna_material = i
        if valor == 'Ce.coste':
            columna_centro_de_coste = i
        if valor == 'Cantidad':
            columna_cantidad = i
    for i in range(1, maxima_fila + 1):
        if hoja.cell(row=i, column=columna_movimiento).value == 'Y15' and hoja.cell(row=i,column=columna_centro_de_coste).value == '12RP151010':
            for casos in recursos:
                for rec in casos:
                    if hoja.cell(row=i, column=columna_material).value in rec:
                        recurso = hoja.cell(row=i, column=columna_material).value
                        vale = hoja.cell(row=i, column=columna_vale).value
                        cantidad = -(hoja.cell(row=i, column=columna_cantidad).value)
                        cantvale = str(vale) + ':' + str(cantidad)
                        if recurso not in vales_por_recurso:
                            vales_por_recurso.update({recurso: [cantvale]})
                        else:
                            lista = vales_por_recurso[recurso]
                            if cantvale not in lista:
                                lista.append(cantvale)
                            vales_por_recurso.update({recurso: lista})
    return vales_por_recurso

def procesa_deficitarios(rec_por_cuenta, recursos_con_vales):
    solucion_optima = False
    total_por_recurso = {}
    for recurso in recursos_con_vales.keys():
        for caso in rec_por_cuenta:
            for orden, rec in enumerate(caso):
                if rec == recurso:
                    cantidad = int(caso[orden + 1])
                    if recurso not in total_por_recurso:
                        total_por_recurso.update({recurso: cantidad})
                    else:
                        existente = int(total_por_recurso[recurso])
                        actual = existente + cantidad
                        total_por_recurso.update({recurso: actual})
    #print(rec_por_cuenta)
    #print(total_por_recurso)
    #print(recursos_con_vales)
    for clave, valor in total_por_recurso.items():
        extraccion = 0
        lista_de_vales = recursos_con_vales[clave]
        for vale in lista_de_vales:
            #print('Los siguientes vales son obligatorios', vale.split(':')[0])
            extraccion += int(vale.split(':')[1])
        #print('Del recurso', clave, 'se usaron', valor, 'y se sacaron de almacen', extraccion)
    for caso in rec_por_cuenta:
        for orden, recurso in enumerate(caso):
            if recurso in recursos_con_vales:
                lista_de_vales = recursos_con_vales[recurso]
                cantidad_a_devolver = int(caso[orden + 1])
                for x in lista_de_vales:
                    cantidad_en_vale = int(x.split(':')[1])
                    if cantidad_a_devolver == cantidad_en_vale:
                        solucion_optima = True
                        if orden == 0:
                            print('Devolucion para Preventivo', x.split(':')[0], recurso, cantidad_en_vale)
                        if orden == 2:
                            print('Devolucion para Inversiones', x.split(':')[0], recurso, cantidad_en_vale)
    devolucion_por_preventivo = []
    devolucion_por_inversiones = []
    nueva_lista = []
    if solucion_optima:
        pass
    else:
        for caso in rec_por_cuenta:
            for orden, recurso in enumerate(caso):
                if recurso in recursos_con_vales:
                    lista_de_vales = recursos_con_vales[recurso]
                    cantidad_a_devolver = int(caso[orden + 1])
                    for i in range(len(lista_de_vales)):
                        if cantidad_a_devolver > 0:
                            if lista_de_vales:
                                x = lista_de_vales.pop(0)
                                vale = x.split(':')[0]
                                cantidad_en_vale = int(x.split(':')[1])
                                if cantidad_a_devolver > cantidad_en_vale:
                                    cantidad_a_devolver -= cantidad_en_vale
                                    remanente_en_vale = 0
                                    if cantidad_en_vale > 0:
                                        guardar = (int(vale), recurso, cantidad_en_vale)
                                        if orden == 0:
                                            devolucion_por_preventivo.append(guardar)
                                        if orden == 2:
                                            devolucion_por_inversiones.append(guardar)
                                else:
                                    guardar = (int(vale), recurso, cantidad_a_devolver)
                                    if orden == 0:
                                        devolucion_por_preventivo.append(guardar)
                                    if orden == 2:
                                        devolucion_por_inversiones.append(guardar)
                                    remanente_en_vale = cantidad_en_vale - cantidad_a_devolver
                                    cantidad_a_devolver = 0
                                    if remanente_en_vale < 0:
                                        remanente_en_vale = 0
                                actualizacion = vale + ':' + str(remanente_en_vale)
                                nueva_lista.append(actualizacion)
                recursos_con_vales.update({recurso: nueva_lista})
    #print(devolucion_por_preventivo)
    #print(devolucion_por_inversiones)
    respuesta = []
    respuesta.append(devolucion_por_preventivo)
    respuesta.append(devolucion_por_inversiones)
    return respuesta



def cantidades_que_no_completan_devolucion_en_un_vale(lista, cantidad_necesaria, vales_todos, recurso):
    devolver = []
    s = list(lista)
    todos = itertools.chain.from_iterable(itertools.combinations(s, r) for r in range(2, len(s) + 1))
    for combinacion in todos:
        temporal = ''
        cantidad_acumulada = 0
        for vale in combinacion:
            lista_de_recursos = vales_todos[int(vale)]
            indice = lista_de_recursos.index(recurso) + 1
            cantidad_en_vale = lista_de_recursos[indice]
            if cantidad_acumulada + cantidad_en_vale <= int(cantidad_necesaria):
                cantidad_acumulada += cantidad_en_vale
                if not temporal:
                    temporal = str(vale) + ':' + str(cantidad_en_vale)
                else:
                    temporal = temporal + '+' + str(vale) + ':' + str(cantidad_en_vale)
            else:
                temporal = temporal + '+' + str(vale) + ':' + str(int(cantidad_necesaria) - cantidad_acumulada)
                devolver.append(temporal)
                break
    return devolver

def funcion_validacion(lista, valor):
    cantidad_consumida = 0
    respuesta = {}
    salir = False
    while not salir:
        entrada = str(input("Indique su elección:\n"))
        if '-' in entrada:
            eleccion = entrada.split('-')[0]
            cantidad = entrada.split('-')[1]
            if eleccion:
                if int(eleccion) in range(len(lista)):
                    if int(cantidad) <= valor:
                        if int(cantidad) + cantidad_consumida <= valor:
                            recurso = lista[int(eleccion)]
                            cantidad_consumida += int(cantidad)
                            respuesta.update({recurso: cantidad})
                            if cantidad_consumida == valor:
                                salir = True
                        else:
                            print('Con la cantidad especificada excede el valor a devolver')
                    else:
                        print('Esta cantidad excede el monto a devolver')
                else:
                    print('Debe especificar un indice para el recurso de los propuestos arriba')
            else:
                print('Debe especificar una elección delante del -')
        else:
            print('Debe separar el indice del recurso de la cantidad por el simbolo -')
    return respuesta


def localiza_vales(recursos, hoja, titulo_columna, devuelve):
    maxima_fila = hoja.max_row
    maxima_columna = hoja.max_column
    vales_por_recurso = {}
    cantidad_de_recurso_en_cada_vale = {}
    for i in range(1, maxima_columna + 1):
        valor = hoja.cell(row=titulo_columna, column=i).value
        if valor == 'CMv':
            columna_movimiento = i
        if valor == 'Doc.mat.':
            columna_vale = i
        if valor == 'Texto breve de material':
            columna_material = i
        if valor == 'Ce.coste':
            columna_centro_de_coste = i
        if valor == 'Cantidad':
            columna_cantidad = i
    for i in range(1, maxima_fila + 1):
        if hoja.cell(row=i, column=columna_movimiento).value == 'Y15' and hoja.cell(row=i, column=columna_centro_de_coste).value == '12RP151010':
            if hoja.cell(row=i, column=columna_material).value in recursos:
                recurso = hoja.cell(row=i, column=columna_material).value
                cantidad = -(hoja.cell(row=i, column=columna_cantidad).value)
                vale = hoja.cell(row=i, column=columna_vale).value
                if vale not in cantidad_de_recurso_en_cada_vale:
                    cantidad_de_recurso_en_cada_vale.update({vale: [recurso, cantidad]})
                else:
                    lista = cantidad_de_recurso_en_cada_vale[vale]
                    lista.append(recurso)
                    lista.append(cantidad)
                    cantidad_de_recurso_en_cada_vale.update({vale: lista})
                if cantidad >= int(recursos[recurso]):
                    if recurso not in vales_por_recurso:
                        vales_por_recurso.update({recurso: [vale]})
                    else:
                        lista = vales_por_recurso[recurso]
                        lista.append(vale)
                        vales_por_recurso.update({recurso: lista})
    for clave in recursos.keys():
        if clave not in vales_por_recurso:
            for i in range(1, maxima_fila + 1):
                if hoja.cell(row=i, column=columna_movimiento).value == 'Y15' and hoja.cell(row=i,column=columna_centro_de_coste).value == '12RP151010':
                    if hoja.cell(row=i, column=columna_material).value == clave:
                        recurso = hoja.cell(row=i, column=columna_material).value
                        vale = hoja.cell(row=i, column=columna_vale).value
                        if recurso not in vales_por_recurso:
                            vales_por_recurso.update({recurso: vale})
                        else:
                            valor_a_actualizar = vales_por_recurso[recurso]
                            valor_a_actualizar = str(valor_a_actualizar) + '+' + str(vale)
                            vales_por_recurso.update({recurso: valor_a_actualizar})
    if 'vales' in devuelve:
        return vales_por_recurso
    else:
        return cantidad_de_recurso_en_cada_vale

def genera_combinaciones_posibles_de_vales(rec_dev_prev, rec_dev_inv, rec_dev_ini, vale_prev, vale_inv, vale_inicio, vales_todos, dist_defic):
    todas_las_elecciones = []
    recursos_ordenados = []
    vales_obligatorios = []
    vales_obligatorios_preventivo = []
    vales_obligatorios_inversiones = []
    vales_obligatorios_devolver_inicio_mes = []
    for clave, valor in vale_prev.items():
        if '+' not in str(valor):
            todas_las_elecciones.append(valor)
            recursos_ordenados.append(clave)
        else:
            cantidad_a_devolver = rec_dev_prev[clave]
            variante = valor.split('+')
            vales_obligatorios_preventivo = cantidades_que_no_completan_devolucion_en_un_vale(variante, cantidad_a_devolver, vales_todos, clave)
    for clave, valor in vale_inv.items():
        if '+' not in valor:
            todas_las_elecciones.append(valor)
            recursos_ordenados.append(clave)
        else:
            cantidad_a_devolver = rec_dev_inv[clave]
            variante = valor.split('+')
            vales_obligatorios_inversiones = cantidades_que_no_completan_devolucion_en_un_vale(variante, cantidad_a_devolver, vales_todos, clave)
    for clave, valor in vale_inicio.items():
        if '+' not in valor:
            todas_las_elecciones.append(valor)
        else:
            cantidad_a_devolver = rec_dev_ini[clave]
            variante = valor.split('+')
            vales_obligatorios_devolver_inicio_mes = cantidades_que_no_completan_devolucion_en_un_vale(variante, cantidad_a_devolver, vales_todos, clave)
    if vales_obligatorios_preventivo:
        vales_obligatorios.append(vales_obligatorios_preventivo)
    if vales_obligatorios_inversiones:
        vales_obligatorios.append(vales_obligatorios_inversiones)
    if vales_obligatorios_devolver_inicio_mes:
        vales_obligatorios.append(vales_obligatorios_devolver_inicio_mes)
    obligados = []
    if vales_obligatorios:
        for caso in vales_obligatorios:
           for cadena in caso:
               lista = cadena.split('+')
               for x in lista:
                    obligados.append(x.split(':')[0])
    respuesta = {}
    combinacion_de_vales = ()
    for combinacion in itertools.product(*todas_las_elecciones):
        variante = elige_combinaciones_de_vales(combinacion, rec_dev_prev, rec_dev_inv, rec_dev_ini, vales_todos, recursos_ordenados)
        if variante is not None:
            if len(variante) > 0:
                if len(respuesta) > 0:
                    for obligatorio in obligados:
                        variante.update({obligatorio: ''})
                    if len(variante) <= len(respuesta):
                        respuesta = variante
                        combinacion_de_vales = combinacion
                else:
                    combinacion_de_vales = combinacion
                    respuesta = variante
                    for obligatorio in obligados:
                        respuesta.update({obligatorio: ''})
    fecha = validacion_de_fecha()
    defic_prev = dist_defic[0]
    defic_inv = dist_defic[1]
    y = list(combinacion_de_vales)
    #print('______________________________________________________________')
    #print('Devolución de recurso para gasto por Mantenimiento Preventivo:')
    imprime_resultado(vale_prev, rec_dev_prev, y, vales_obligatorios_preventivo, 'preventivo', fecha, defic_prev)
    #print('______________________________________________________________'),
    #print('Devolución de recurso para gasto por Inversiones:')
    imprime_resultado(vale_inv, rec_dev_inv, y, vales_obligatorios_inversiones, 'inversiones', fecha, defic_inv)
    #print('______________________________________________________________')
    #print('Devolución de recurso para gasto para Sacar a inicio de mes:')
    imprime_resultado(vale_inicio, rec_dev_ini, y, vales_obligatorios_devolver_inicio_mes, 'inicio', fecha, [])

def validacion_de_fecha():
    year = '2022'
    mes_ultimo_dia = {1: 31, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}
    salir = False
    posibles = [str(i) for i in range(1, 13)]
    febrero = ['28', '29']
    print('Indique el mes que se analiza para el cierre')
    while not salir:
        entrada = str(input("Entre un numero entre 1 y 12:\n"))
        if entrada in posibles:
            if entrada == '2':
                salir_de_febrero = False
                while not salir_de_febrero:
                    print("Especifique el ultimo dia de febrero:\n")
                    dia = input('Escriba 28 0 29:\n')
                    if dia in febrero:
                        salir_de_febrero = True
                        salir = True
            else:
                dia = str(mes_ultimo_dia[int(entrada)])
                salir = True
    if int(entrada) < 10:
        mes = '0' + str(entrada)
    return str(year + '/' + mes + '/' + dia)


def imprime_resultado(recursos_con_vales, recursos_con_cantidad, combinacion, vales_obligatorios, tipo_de_cuenta, fecha, distribuc_defic):
    if tipo_de_cuenta == 'preventivo' or tipo_de_cuenta == 'inversiones':
        observaciones = 'CAMBIO DE CUENTA según vale de salida'
    else:
        observaciones = 'POR NO UTILIZACION en el mes según vale de salida'
    excel_con_datos = glob.glob('*.xlsx')
    contador = 0
    for libro in excel_con_datos:
        if 'Modelo de Logistica Inv.xlsx' in libro:
            contador += 1
            logistica_inversa = libro
    if contador == 1:
        #print('Se procesará el siguiente archivo de Logística Inversa:', logistica_inversa)
        wb = openpyxl.load_workbook(logistica_inversa)
        ws1 = wb.sheetnames
        for hoja in ws1:
            if hoja == 'LOG_INV':
                sheet = wb[hoja]
        for i in range(2, 13):
            if sheet.cell(row=2, column=i).value == 'CONSECUTIVO':
                indice_consecutivo = i
            if sheet.cell(row=2, column=i).value == 'FECHA':
                indice_fecha = i
            if sheet.cell(row=2, column=i).value == 'DESCRIPCION':
                indice_descripcion = i
            if sheet.cell(row=2, column=i).value == 'CANTIDAD':
                indice_cantidad = i
            if sheet.cell(row=2, column=i).value == 'OBSERVACIONES':
                indice_observaciones = i
            if sheet.cell(row=2, column=i).value == 'VALE ENTRADA':
                indice_vale = i
        primeros_cuatro = str(fecha[2:4] + fecha[5:7])
        for i in range(3, 3600):
            if i == 3 and sheet.cell(row=i, column=2).value is None:
                consecutivo = 1
            if i > 3 and sheet.cell(row=i, column=2).value is None:
                anterior = str(sheet.cell(row=i - 1, column=2).value)
                if fecha[5:7] == anterior[2:4]:
                    consecutivo = int(anterior[-2:]) + 1
                else:
                    consecutivo = 1
            if sheet.cell(row=i, column=7).value is None:
                comienzo_a_escribir = i
                break
        orden_de_obligatorios = -1
        if combinacion:
            tuplas_vales = []
            for recurso, sus_vales in recursos_con_vales.items():
                cantidad = recursos_con_cantidad[recurso]
                if combinacion:
                    print_vale = combinacion.pop(0)
                    if '+' not in sus_vales:
                        tuplas_vales.append((print_vale, recurso, cantidad))
                        #print(print_vale, recurso, cantidad)
                    else:
                        orden_de_obligatorios += 1
                        lista = vales_obligatorios[orden_de_obligatorios].split('+')
                        for vale in lista:
                            print_vale = vale.split(':')[0]
                            cantidad = vale.split(':')[1]
                            tuplas_vales.append((print_vale, recurso, cantidad))
            tuplas_vales += distribuc_defic
            ordenados = sorted(tuplas_vales, key=lambda x : x[0])
            x = -1
            vale_actual = ''
            incrementar_consecutivo = -1
            for caso in ordenados:
                if caso[0] != vale_actual:
                    incrementar_consecutivo += 1
                    vale_actual = caso[0]
                x += 1
                ultimos_dos = consecutivo + incrementar_consecutivo
                if ultimos_dos < 10:
                    ultimos_dos = '0' + str(ultimos_dos)
                else:
                    ultimos_dos = str(ultimos_dos)
                sheet.cell(row=comienzo_a_escribir + x, column=indice_vale).value = caso[0]
                sheet.cell(row=comienzo_a_escribir + x, column=indice_descripcion).value = caso[1]
                sheet.cell(row=comienzo_a_escribir + x, column=indice_cantidad).value = caso[2]
                sheet.cell(row=comienzo_a_escribir + x, column=indice_observaciones).value = observaciones
                sheet.cell(row=comienzo_a_escribir + x, column=indice_fecha).value = fecha
                sheet.cell(row=comienzo_a_escribir + x, column=indice_consecutivo).value = int(primeros_cuatro + ultimos_dos)
    wb.save(logistica_inversa)
    wb.close()


def elige_combinaciones_de_vales(combinacion, rec_dev_prev, rec_dev_inv, rec_dev_ini, vales_todos, recursos_ordenados):
    lista_de_combinaciones = list(combinacion)
    respuesta = {}
    for comb in lista_de_combinaciones:
        respuesta.update({comb: ''})
    recursos_a_devolver_por_preventivo = copy.deepcopy(rec_dev_prev)
    recursos_a_devolver_por_inversiones = copy.deepcopy(rec_dev_inv)
    recursos_a_devolver_inicio = copy.deepcopy(rec_dev_ini)
    vales_contra_recursos = copy.deepcopy(vales_todos)
    sirve = True
    for recurso in recursos_ordenados:
        vale = lista_de_combinaciones.pop(0)
        if recurso in recursos_a_devolver_por_preventivo:
            cantidad_a_devolver = int(recursos_a_devolver_por_preventivo[recurso])
            del recursos_a_devolver_por_preventivo[recurso]
        else:
            if recurso in recursos_a_devolver_por_inversiones:
                cantidad_a_devolver = int(recursos_a_devolver_por_inversiones[recurso])
                del recursos_a_devolver_por_inversiones[recurso]
            else:
                if recurso in recursos_a_devolver_inicio:
                    cantidad_a_devolver = int(recursos_a_devolver_inicio[recurso])
                    del recursos_a_devolver_inicio[recurso]
        lista_de_recursos = vales_contra_recursos[vale]
        if recurso in lista_de_recursos:
            indice = lista_de_recursos.index(recurso) + 1
            cantidad_en_existencia = lista_de_recursos[indice]
            if int(cantidad_en_existencia) >= cantidad_a_devolver:
                cantidad_en_existencia -= int(cantidad_a_devolver)
                lista_de_recursos[indice] = int(cantidad_en_existencia)
                vales_contra_recursos.update({vale: lista_de_recursos})
            else:
                sirve = False
                break
    if sirve:
        return respuesta
    else:
        respuesta = {}
        return respuesta

excel_con_datos = glob.glob('*.xlsx')
cambio_de_cuenta_preventivo = {}
cambio_de_cuenta_inversiones = {}
sacar_al_correctivo_inicio_mes = {}
contador = 0
for libro in excel_con_datos:
    if 'Resultados' in libro:
        contador += 1
        libro_resultado = libro
if contador == 1:
    print('Se procesará el siguiente archivo de Resultados:', libro_resultado)
    wb = openpyxl.load_workbook(libro_resultado)
    ws1 = wb.sheetnames
    print(ws1)
    for hoja in ws1:
        if hoja == 'Devoluciones':
            sheet = wb[hoja]
    if sheet:
        if sheet.cell(row=1, column=1).value == 'Descriptivo':
            for i in range(2, 13):
                if sheet.cell(row=i, column=1).value != None and sheet.cell(row=i, column=1).value != 'Total general':
                    descriptivo = sheet.cell(row=i, column=1).value
                    if descriptivo != 'Grapa plastica':
                        valor_preventivo = sheet.cell(row=i, column=2).value
                        valor_inversiones = sheet.cell(row=i, column=3).value
                        valor_sacar = sheet.cell(row=i, column=4).value
                    else:
                        valor_preventivo = int(sheet.cell(row=i, column=2).value)/100
                        valor_inversiones = int(sheet.cell(row=i, column=3).value)/100
                        valor_sacar = sheet.cell(row=i, column=4).value
                    cambio_de_cuenta_preventivo.update({descriptivo: valor_preventivo})
                    cambio_de_cuenta_inversiones.update({descriptivo: valor_inversiones})
                    sacar_al_correctivo_inicio_mes.update({descriptivo: valor_sacar})
                else:
                    break
    else:
        print("ADVERTENCIA: Fue eliminada la hoja Devoluciones")
    wb.close()

descriptivo_con_lista_de_materiales = {}
recurso_devolver_preventivo = {}
recurso_devolver_inversiones = {}
recurso_devolver_inicio_mes = {}
contador = 0
for libro in excel_con_datos:
    if 'Vales' in libro:
        contador += 1
        vales = libro
if contador == 1:
    print('Se procesará el siguiente archivo de Vales:', vales)
    wb = openpyxl.load_workbook(vales)
    sheet = wb.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    for i in range(1, max_row + 1):
        if sheet.cell(row=i, column=max_column).value != None:
            indice_de_titulos = i
            break
    nombres_de_columnas = []
    for j in range(1, max_column + 1):
        nombres_de_columnas.append(sheet.cell(row=indice_de_titulos, column=j).value)
    for indice, columna in enumerate(nombres_de_columnas):
        if 'Descriptivo' in columna:
            indice_columna_descriptivo = indice + 1
        if 'Texto breve de material' in columna:
            indice_columna_texto_material = indice + 1
        if 'CMv' in columna:
            indice_columna_movimiento = indice + 1
        if 'Ce.coste' in columna:
            indice_columna_centro_de_coste = indice + 1
    for i in range(1, max_row + 1):
        if sheet.cell(row=i, column=indice_columna_descriptivo).value != None:
            if sheet.cell(row=i, column=indice_columna_movimiento).value == 'Y15':
                descriptivo = sheet.cell(row=i, column=indice_columna_descriptivo).value
                material = sheet.cell(row=i, column=indice_columna_texto_material).value
                if descriptivo not in descriptivo_con_lista_de_materiales:
                    descriptivo_con_lista_de_materiales.update({descriptivo:[]})
                else:
                    lista_existente = descriptivo_con_lista_de_materiales[descriptivo]
                    if material not in lista_existente:
                        lista_existente.append(material)
                        descriptivo_con_lista_de_materiales.update({descriptivo: lista_existente})
    print('A continuación se le presentará la cantidad de cada recurso a devolver por cada concepto para que elija el monto por cada elemento específico')
    print('Use el siguiente convenio: "orden_del_recurso"-"cantidad"')
    print('+++++++++++++++++++++++++++++++++++++++++++++++ CAMBIO DE CUENTA PARA PREVENTIVO: +++++++++++++++++++++++++++++++++++++++++++++++')
    for clave, valor in cambio_de_cuenta_preventivo.items():
        if valor > 0:
            lista_materiales = descriptivo_con_lista_de_materiales[clave]
            if len(lista_materiales) > 1:
                print('Del recurso', clave, 'se deben devolver:', valor)
                for orden, elemento in enumerate(lista_materiales):
                    print(orden, '-', elemento)
                recurso_devolver_preventivo.update(funcion_validacion(lista_materiales, valor))
                print('--------------------------------------------------')
            else:
                recurso_devolver_preventivo.update({lista_materiales[0]: valor})
    print('+++++++++++++++++++++++++++++++++++++++++++++++ CAMBIO DE CUENTA PARA INVERSIONES: +++++++++++++++++++++++++++++++++++++++++++++++')
    for clave, valor in cambio_de_cuenta_inversiones.items():
        if valor > 0:
            lista_materiales = descriptivo_con_lista_de_materiales[clave]
            if len(lista_materiales) > 1:
                print('Del recurso', clave, 'se deben devolver:', valor)
                for orden, elemento in enumerate(lista_materiales):
                    print(orden, '-', elemento)
                recurso_devolver_inversiones.update(funcion_validacion(lista_materiales, valor))
                print('--------------------------------------------------')
            else:
                recurso_devolver_inversiones.update({lista_materiales[0]: valor})
    print('+++++++++++++++++++++++++++++++++++++++++++++++ DEVOLUCION INICIO DE MES: +++++++++++++++++++++++++++++++++++++++++++++++')
    recursos_con_devolucion = []
    for clave, valor in sacar_al_correctivo_inicio_mes.items():
        if valor > 0:
            lista_materiales = descriptivo_con_lista_de_materiales[clave]
            if len(lista_materiales) > 1:
                print('Del recurso', clave, 'se deben devolver:', valor)
                for orden, elemento in enumerate(lista_materiales):
                    print(orden, '-', elemento)
                recurso_devolver_inicio_mes.update(funcion_validacion(lista_materiales, valor))
                print('--------------------------------------------------')
            else:
                recurso_devolver_inicio_mes.update({lista_materiales[0]: valor})
        elif valor < 0:
            lista_materiales = descriptivo_con_lista_de_materiales[clave]
            recursos_con_devolucion.append(lista_materiales)
            print('********************************************** IMPORTANTE **********************************************')
            print('Se le debe dar entrada al siguiente recurso recuperado de la Operación en el mes actual para igualar el consumo con las extracciones:', clave, valor)
            print('********************************************************************************************************')
    rec_dev = []
    for deficitario in recursos_con_devolucion:
        for caso in deficitario:
            lista = []
            if caso in recurso_devolver_preventivo:
                lista.append(caso)
                lista.append(recurso_devolver_preventivo[caso])
                del recurso_devolver_preventivo[caso]
            if caso in recurso_devolver_inversiones:
                lista.append(caso)
                lista.append(recurso_devolver_inversiones[caso])
                del recurso_devolver_inversiones[caso]
            if lista:
                rec_dev.append(lista)
    #print(rec_dev)
    #print(recurso_devolver_preventivo)
    #print(recurso_devolver_inversiones)
    #print(recurso_devolver_inicio_mes)
    todos_los_recursos = {}
    todos_los_recursos.update(recurso_devolver_preventivo)
    todos_los_recursos.update(recurso_devolver_inversiones)
    todos_los_recursos.update(recurso_devolver_inicio_mes)
    #print('--------------------------------------------------------------------------------------')
    vales_para_preventivo = localiza_vales(recurso_devolver_preventivo, sheet, indice_de_titulos, 'vales')
    #for clave, valor in vales_para_preventivo.items():
    #    print(clave, valor)
    #print('--------------------------------------------------------------------------------------')
    vales_para_inversiones = localiza_vales(recurso_devolver_inversiones, sheet, indice_de_titulos, 'vales')
    #for clave, valor in vales_para_inversiones.items():
    #    print(clave, valor)
    #print('--------------------------------------------------------------------------------------')
    vales_para_inicio_de_mes = localiza_vales(recurso_devolver_inicio_mes, sheet, indice_de_titulos, 'vales')
    #for clave, valor in vales_para_inicio_de_mes.items():
    #    print(clave, valor)
    #print('--------------------------------------------------------------------------------------')
    todos_los_vales = localiza_vales(todos_los_recursos, sheet, indice_de_titulos, 'todos')
    #for clave, valor in todos_los_vales.items():
    #    print(clave, valor)
    vales_de_deficitarios = analisis_de_recursos_donde_consumo_excede_salida_almacen(rec_dev, sheet, indice_de_titulos)
    #print('Vales de deficitarios:', vales_de_deficitarios)
    deficitarios  = procesa_deficitarios(rec_dev, vales_de_deficitarios)
    genera_combinaciones_posibles_de_vales(recurso_devolver_preventivo, recurso_devolver_inversiones, recurso_devolver_inicio_mes, vales_para_preventivo, vales_para_inversiones, vales_para_inicio_de_mes, todos_los_vales, deficitarios)
    wb.close()

end = timeit.timeit()
print(end - start)








